#!/usr/bin/env python3
"""Regenerate outputs using original Excel templates (layout-preserving)."""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import openpyxl

from etl_load_sources import infer_region_id, norm, norm_key

RAW_SOURCES = [
    "files/7-MTL for Table and DailySales(2026 Jan to Mar).xlsx",
    "files/9-MLM Table and DailySales-2026_Feb.xlsx",
    "files/MHL 2026 Feb.xlsx",
    "files/Keng Tung - Jan New Update.xlsx",
    "files/LSH Feb PJP.xlsx",
]

MEASURE_TOKENS = {"pkt", "pk", "pack", "bot", "bottle", "lit", "liter", "litre"}

MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


EXPORT_CATEGORIES = {
    "individual_sales",
    "sku_summary",
    "township_summary",
    "van_wise_sku",
    "sales_compare",
    "follow_up_sales",
    "debtors",
    "other",
}


def classify_template(filename: str) -> str:
    name = norm_key(filename)
    if "individual" in name:
        return "individual_sales"
    if "van wise" in name:
        return "van_wise_sku"
    if "sku summary" in name or "sku wise" in name or "sku analysis" in name:
        return "sku_summary"
    if "township summary" in name:
        return "township_summary"
    if "sales compare" in name or "compare for" in name or name.startswith("compare"):
        return "sales_compare"
    if "follow up" in name or "followup" in name or "fus" in name:
        return "follow_up_sales"
    if "debtor" in name:
        return "debtors"
    return "other"


def normalize_filename(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", name.lower()).strip()


def core_input_reason(filename: str) -> str | None:
    key = normalize_filename(filename)
    if "table" in key and ("dailysales" in key or "daily sales" in key):
        return "Table + DailySales"
    if "pg" in key and ("pg sales" in key or "pg daily sales" in key or "pg dailysales" in key):
        return "PG Table + PGSales"
    return None


def prepare_in(root: Path, in_dir: Path) -> List[str]:
    in_dir.mkdir(parents=True, exist_ok=True)
    copied = []
    for rel in RAW_SOURCES:
        src = root / rel
        if not src.exists():
            continue
        dest = in_dir / src.name
        if not dest.exists():
            shutil.copy2(src, dest)
            copied.append(src.name)
    return copied


def read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def to_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if s == "":
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def month_key(date_val, year_val, month_val) -> str:
    if date_val not in (None, ""):
        s = str(date_val)
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                d = dt.datetime.strptime(s[:10], fmt)
                return f"{d.year:04d}-{d.month:02d}"
            except Exception:
                pass
        if isinstance(date_val, (dt.date, dt.datetime)):
            d = date_val
            return f"{d.year:04d}-{d.month:02d}"
    try:
        y = int(float(year_val)) if year_val not in (None, "") else None
        m = int(float(month_val)) if month_val not in (None, "") else None
        if y and m:
            return f"{y:04d}-{m:02d}"
    except Exception:
        return ""
    return ""


def parse_month_label(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, (dt.date, dt.datetime)):
        return f"{val.year:04d}-{val.month:02d}"
    s = str(val).strip()
    if s == "":
        return None
    # YYYY-MM or YYYY/MM
    m = re.search(r"(20\d{2})[\-/](\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    # Month name + year
    m = re.search(r"([A-Za-z]+)[\s\-_/]*(20\d{2})", s)
    if m:
        mon = MONTHS.get(m.group(1).lower())
        if mon:
            return f"{int(m.group(2)):04d}-{mon:02d}"
    # Year + month name
    m = re.search(r"(20\d{2}).*?([A-Za-z]+)", s)
    if m:
        mon = MONTHS.get(m.group(2).lower())
        if mon:
            return f"{int(m.group(1)):04d}-{mon:02d}"
    return None


def parse_month_from_sheet(name: str) -> Optional[str]:
    return parse_month_label(name)


def find_row_with_tokens(ws, tokens: List[str], max_scan: int = 15) -> Tuple[int, Dict[int, str]]:
    best_row = 0
    best_score = 0
    best_map: Dict[int, str] = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        row_vals = [norm_key(v) for v in row]
        score = sum(1 for v in row_vals if v in tokens)
        if score > best_score:
            best_score = score
            best_row = r_idx
            col_map = {}
            for idx, v in enumerate(row_vals, start=1):
                if v in tokens:
                    col_map[idx] = v
            best_map = col_map
    return best_row, best_map


def build_merge_map(ws) -> Dict[Tuple[int, int], Tuple[int, int]]:
    merge_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merge_map[(r, c)] = (rng.min_row, rng.min_col)
    return merge_map


def safe_write(ws, r: int, c: int, value, merge_map: Dict[Tuple[int, int], Tuple[int, int]]):
    anchor = merge_map.get((r, c))
    if anchor and anchor != (r, c):
        # avoid writing into non-anchor merged cells
        return
    cell = ws.cell(r, c)
    # keep formulas intact
    if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
        return
    cell.value = value


def extract_month_columns(ws, measure_row: int, month_row: int) -> List[Tuple[int, str, str]]:
    cols = []
    row_vals = [norm_key(v) for v in next(ws.iter_rows(min_row=measure_row, max_row=measure_row, values_only=True))]
    for c_idx, token in enumerate(row_vals, start=1):
        if token not in MEASURE_TOKENS:
            continue
        # find month label from month_row scanning left
        month_val = None
        for c2 in range(c_idx, 0, -1):
            val = ws.cell(month_row, c2).value
            if val is not None and str(val).strip() != "":
                month_val = val
                break
        month_key_val = parse_month_label(month_val)
        if not month_key_val:
            continue
        cols.append((c_idx, month_key_val, token))
    return cols


def is_total_label(text: str) -> bool:
    t = norm_key(text)
    return t in ("total", "grand total", "subtotal", "totaling") or t.startswith("total")


def normalize_product_key(name: str, ml, packing) -> str:
    return norm_key(f"{name}|{ml}|{packing}")


def load_outlet_aliases(path: Path):
    exact = {}
    name_only = {}
    if not path.exists():
        return exact, name_only
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            an = row.get("alias_name") or ""
            at = row.get("alias_township") or ""
            cn = row.get("canonical_name") or ""
            ct = row.get("canonical_township") or ""
            if not an or not cn:
                continue
            akey = (norm_key(an), norm_key(at))
            cval = (norm_key(cn), norm_key(ct))
            if akey[1]:
                exact[akey] = cval
            else:
                name_only[akey[0]] = cval
    return exact, name_only


def resolve_outlet_key(name, township, exact_map, name_only_map):
    n = norm_key(name)
    t = norm_key(township)
    if (n, t) in exact_map:
        cn, ct = exact_map[(n, t)]
        return cn, ct
    if n in name_only_map:
        cn, ct = name_only_map[n]
        return cn, ct
    return n, t


def load_product_aliases(path: Path):
    exact = {}
    name_only = {}
    if not path.exists():
        return exact, name_only
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            an = row.get("alias_name") or ""
            aml = row.get("alias_ml") or ""
            ap = row.get("alias_packing") or ""
            cn = row.get("canonical_name") or ""
            cml = row.get("canonical_ml") or ""
            cp = row.get("canonical_packing") or ""
            if not an or not cn:
                continue
            akey = norm_key(f"{an}|{aml}|{ap}")
            cval = (cn, cml, cp)
            if aml or ap:
                exact[akey] = cval
            else:
                name_only[norm_key(an)] = cval
    return exact, name_only


def resolve_product_key(name, ml, packing, exact_map, name_only_map):
    key = norm_key(f"{name}|{ml}|{packing}")
    if key in exact_map:
        cn, cml, cp = exact_map[key]
        return normalize_product_key(cn, cml, cp)
    n = norm_key(name)
    if n in name_only_map:
        cn, cml, cp = name_only_map[n]
        return normalize_product_key(cn, cml, cp)
    return normalize_product_key(name, ml, packing)


def simplify_name(text: str) -> str:
    if not text:
        return ""
    s = str(text)
    s = s.replace("×", "x")
    s = re.sub(r"\(.*?\)", "", s)
    s = re.sub(r"\d+(\.\d+)?", " ", s)
    s = re.sub(r"[^0-9A-Za-z\u1000-\u109F]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip().casefold()
    return s


def build_aggregates_by_region(products: List[Dict[str, str]], outlets: List[Dict[str, str]], sales: List[Dict[str, str]]):
    product_by_id = {p.get("product_id"): p for p in products if p.get("product_id")}
    outlet_by_id = {o.get("outlet_id"): o for o in outlets if o.get("outlet_id")}

    aggs: Dict[str, dict] = {}

    def get_agg(region_id: str):
        if region_id not in aggs:
            aggs[region_id] = {
                "ind": defaultdict(lambda: defaultdict(lambda: {"pkt": 0.0, "bot": 0.0, "lit": 0.0})),
                "sku": defaultdict(lambda: defaultdict(lambda: {"bot": 0.0, "lit": 0.0})),
                "town": defaultdict(lambda: defaultdict(lambda: {"bottle": 0.0, "liter": 0.0})),
                "town_prod": defaultdict(lambda: defaultdict(lambda: {"pkt": 0.0, "bottle": 0.0, "liter": 0.0})),
                "van": defaultdict(lambda: {"bottle": 0.0, "liter": 0.0}),
                "months_present": set(),
                "product_by_id": product_by_id,
            }
        return aggs[region_id]

    for row in sales:
        region_id = row.get("region_id") or infer_region_id(row.get("source_file", ""))
        agg = get_agg(region_id)
        row["month_key"] = month_key(row.get("date"), row.get("year"), row.get("month"))
        month = row.get("month_key")
        if not month:
            continue
        agg["months_present"].add(month)

        outlet_name = norm(row.get("outlet_name_raw"))
        township = norm(row.get("township_name_raw"))
        outlet_id = row.get("outlet_id")
        if not outlet_name and outlet_id:
            outlet_name = outlet_by_id.get(outlet_id, {}).get("outlet_name_mm", "")
        if not township and outlet_id:
            township = outlet_by_id.get(outlet_id, {}).get("township_name", "")

        qty_pack = to_float(row.get("qty_pack"))
        qty_bot = to_float(row.get("qty_bottle"))
        qty_lit = to_float(row.get("qty_liter"))

        if outlet_name:
            key = (norm_key(outlet_name), norm_key(township))
            agg["ind"][key][month]["pkt"] += qty_pack
            agg["ind"][key][month]["bot"] += qty_bot
            agg["ind"][key][month]["lit"] += qty_lit

        pid = row.get("product_id")
        prod = product_by_id.get(pid, {}) if pid else {}
        pname = prod.get("product_name") or row.get("stock_name_raw") or ""
        pml = prod.get("ml") or row.get("ml_raw") or ""
        ppack = prod.get("packing") or row.get("packing_raw") or ""
        pkey = normalize_product_key(pname, pml, ppack)
        if pname:
            agg["sku"][pkey][month]["bot"] += qty_bot
            agg["sku"][pkey][month]["lit"] += qty_lit
        if township and pname:
            tkey = (norm_key(township), pkey)
            agg["town_prod"][tkey][month]["pkt"] += qty_pack
            agg["town_prod"][tkey][month]["bottle"] += qty_bot
            agg["town_prod"][tkey][month]["liter"] += qty_lit

        if township:
            tkey = norm_key(township)
            agg["town"][tkey][month]["bottle"] += qty_bot
            agg["town"][tkey][month]["liter"] += qty_lit

        van_id = norm(row.get("car_no"))
        if township and van_id and pname:
            pkey_van = simplify_name(pname)
            if pkey_van:
                vkey = (norm_key(township), norm_key(van_id), month, pkey_van)
                agg["van"][vkey]["bottle"] += qty_bot
                agg["van"][vkey]["liter"] += qty_lit

    return aggs


def fill_individual_sales(ws, agg, months_present: set, outlet_alias_exact, outlet_alias_name):
    # find measure row (pkt/bot/lit)
    measure_row, _ = find_row_with_tokens(ws, ["pkt", "bot", "lit", "liter"], max_scan=10)
    if not measure_row:
        return
    month_row = max(1, measure_row - 1)
    label_row = measure_row

    # find label columns
    label_vals = [norm_key(v) for v in next(ws.iter_rows(min_row=label_row, max_row=label_row, values_only=True))]
    customer_col = None
    township_col = None
    for idx, v in enumerate(label_vals, start=1):
        if v in ("customer name", "customer names", "customer", "ကုန်သည်အမည်"):
            customer_col = idx
        if v in ("township", "area", "township (area)"):
            township_col = idx
    if not customer_col:
        customer_col = 2
    if not township_col:
        township_col = 4

    month_cols = extract_month_columns(ws, measure_row, month_row)

    merge_map = build_merge_map(ws)
    for r in range(measure_row + 1, ws.max_row + 1):
        cust = ws.cell(r, customer_col).value
        town = ws.cell(r, township_col).value
        if is_total_label(str(cust)):
            continue
        if cust is None or str(cust).strip() == "":
            continue
        cn, ct = resolve_outlet_key(cust, town, outlet_alias_exact, outlet_alias_name)
        key = (cn, ct)
        for c_idx, mkey, token in month_cols:
            if mkey not in months_present:
                continue
            measure = "pkt" if token in ("pkt", "pk", "pack") else "bot" if token in ("bot", "bottle") else "lit"
            val = agg.get(key, {}).get(mkey, {}).get(measure, 0.0)
            safe_write(ws, r, c_idx, val, merge_map)


def fill_sku_summary(ws, agg, months_present: set, prod_alias_exact, prod_alias_name):
    measure_row, _ = find_row_with_tokens(ws, ["bot", "lit", "bottle", "liter"], max_scan=10)
    if not measure_row:
        return
    month_row = max(1, measure_row - 1)
    label_row = month_row

    header_vals = [norm_key(v) for v in next(ws.iter_rows(min_row=label_row, max_row=label_row, values_only=True))]
    product_col = None
    ml_col = None
    packing_col = None
    for idx, v in enumerate(header_vals, start=1):
        if v in ("product name", "particular", "product", "stockname"):
            product_col = idx
        if v in ("ml",):
            ml_col = idx
        if v in ("packing", "qty"):
            packing_col = idx
    if not product_col:
        product_col = 2

    month_cols = extract_month_columns(ws, measure_row, month_row)

    merge_map = build_merge_map(ws)
    for r in range(measure_row + 1, ws.max_row + 1):
        pname = ws.cell(r, product_col).value
        if is_total_label(str(pname)):
            continue
        if pname is None or str(pname).strip() == "":
            continue
        pml = ws.cell(r, ml_col).value if ml_col else ""
        pack = ws.cell(r, packing_col).value if packing_col else ""
        key = resolve_product_key(pname, pml, pack, prod_alias_exact, prod_alias_name)
        for c_idx, mkey, token in month_cols:
            if mkey not in months_present:
                continue
            measure = "bot" if token in ("bot", "bottle") else "lit"
            val = agg.get(key, {}).get(mkey, {}).get(measure, 0.0)
            safe_write(ws, r, c_idx, val, merge_map)


def fill_township_summary(ws, agg, months_present: set, outlet_alias_exact, outlet_alias_name):
    measure_row, _ = find_row_with_tokens(ws, ["bottle", "liter", "bot", "lit"], max_scan=10)
    if not measure_row:
        return
    month_row = max(1, measure_row - 1)
    label_row = month_row

    header_vals = [norm_key(v) for v in next(ws.iter_rows(min_row=label_row, max_row=label_row, values_only=True))]
    township_col = None
    for idx, v in enumerate(header_vals, start=1):
        if v in ("township", "township( area)", "township (area)"):
            township_col = idx
    if not township_col:
        township_col = 2

    month_cols = extract_month_columns(ws, measure_row, month_row)

    merge_map = build_merge_map(ws)
    for r in range(measure_row + 1, ws.max_row + 1):
        town = ws.cell(r, township_col).value
        if is_total_label(str(town)):
            continue
        if town is None or str(town).strip() == "":
            continue
        _, ct = resolve_outlet_key("", town, outlet_alias_exact, outlet_alias_name)
        tkey = ct or norm_key(town)
        for c_idx, mkey, token in month_cols:
            if mkey not in months_present:
                continue
            measure = "bottle" if token in ("bot", "bottle") else "liter"
            val = agg.get(tkey, {}).get(mkey, {}).get(measure, 0.0)
            safe_write(ws, r, c_idx, val, merge_map)


def fill_township_detail(ws, agg, township_name: str, months_present: set, prod_alias_exact, prod_alias_name):
    measure_row, _ = find_row_with_tokens(ws, ["pk", "pkt", "bottle", "liter", "bot", "lit"], max_scan=10)
    if not measure_row:
        return
    month_row = max(1, measure_row - 1)
    label_row = month_row

    header_vals = [norm_key(v) for v in next(ws.iter_rows(min_row=label_row, max_row=label_row, values_only=True))]
    product_col = None
    ml_col = None
    packing_col = None
    for idx, v in enumerate(header_vals, start=1):
        if v in ("product name", "particular", "product", "stockname"):
            product_col = idx
        if v in ("ml",):
            ml_col = idx
        if v in ("packing", "qty"):
            packing_col = idx
    if not product_col:
        product_col = 2

    month_cols = extract_month_columns(ws, measure_row, month_row)
    tkey = norm_key(township_name)

    merge_map = build_merge_map(ws)
    for r in range(measure_row + 1, ws.max_row + 1):
        pname = ws.cell(r, product_col).value
        if is_total_label(str(pname)):
            continue
        if pname is None or str(pname).strip() == "":
            continue
        pml = ws.cell(r, ml_col).value if ml_col else ""
        pack = ws.cell(r, packing_col).value if packing_col else ""
        pkey = resolve_product_key(pname, pml, pack, prod_alias_exact, prod_alias_name)
        key = (tkey, pkey)
        for c_idx, mkey, token in month_cols:
            if mkey not in months_present:
                continue
            if token in ("pkt", "pk", "pack"):
                measure = "pkt"
            elif token in ("bot", "bottle"):
                measure = "bottle"
            else:
                measure = "liter"
            val = agg.get(key, {}).get(mkey, {}).get(measure, 0.0)
            safe_write(ws, r, c_idx, val, merge_map)


def fill_van_wise(ws, van_agg, months_present: set, prod_alias_exact, prod_alias_name):
    # headers in row 2
    header_row = 2
    header_vals = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    # identify product columns
    product_cols = []  # list of (product_name, bottle_col, liter_col)
    c = 1
    while c <= len(header_vals):
        val = header_vals[c - 1]
        vnorm = norm_key(val)
        if vnorm in ("region", "township", "date") or vnorm in ("bottle", "liter") or val is None:
            c += 1
            continue
        # treat as product label, expect Bottle/Liter next
        bottle_col = c + 1 if c + 1 <= len(header_vals) and norm_key(header_vals[c]) in ("bottle", "bot") else None
        liter_col = c + 2 if c + 2 <= len(header_vals) and norm_key(header_vals[c + 1]) in ("liter", "lit") else None
        if bottle_col and liter_col:
            product_cols.append((str(val), bottle_col, liter_col))
            c += 3
        else:
            c += 1

    sheet_month = parse_month_from_sheet(ws.title)
    if not sheet_month or sheet_month not in months_present:
        return

    # data rows start at row 5
    merge_map = build_merge_map(ws)
    for r in range(5, ws.max_row + 1):
        town = ws.cell(r, 2).value
        van = ws.cell(r, 3).value
        if town is None or str(town).strip() == "":
            continue
        tkey = norm_key(town)
        vkey = norm_key(van)
        for pname, bcol, lcol in product_cols:
            # resolve alias by name (ignore ml/packing in van-wise headers)
            canon_name = pname
            nkey = norm_key(pname)
            if nkey in prod_alias_name:
                canon_name = prod_alias_name[nkey][0]
            pkey = simplify_name(canon_name)
            key = (tkey, vkey, sheet_month, pkey)
            vals = van_agg.get(key)
            if not vals:
                safe_write(ws, r, bcol, 0.0, merge_map)
                safe_write(ws, r, lcol, 0.0, merge_map)
            else:
                safe_write(ws, r, bcol, vals.get("bottle", 0.0), merge_map)
                safe_write(ws, r, lcol, vals.get("liter", 0.0), merge_map)


def regenerate_templates(
    root: Path,
    in_dir: Path,
    out_dir: Path,
    clean_out: bool = True,
    include: Optional[List[str]] = None,
    regions: Optional[List[str]] = None,
    include_map: Optional[Dict[str, List[str]]] = None,
    include_files_map: Optional[Dict[str, List[str]]] = None,
):
    staging_dir = out_dir / "staging"
    include_set = set(i for i in (include or []) if i)
    region_filter = {r.upper() for r in regions} if regions else None

    def pick_master(name: str, fallback: str) -> List[Dict[str, str]]:
        master = staging_dir / name
        if master.exists():
            return read_csv(master)
        return read_csv(staging_dir / fallback)

    products = pick_master("master_products.csv", "products.csv")
    outlets = pick_master("master_outlets.csv", "outlets.csv")
    sales = pick_master("master_sales_transactions.csv", "sales_transactions.csv")

    aggs = build_aggregates_by_region(products, outlets, sales)

    outlet_alias_exact, outlet_alias_name = load_outlet_aliases(staging_dir / "aliases_outlets.csv")
    prod_alias_exact, prod_alias_name = load_product_aliases(staging_dir / "aliases_products.csv")

    # per region output
    template_root = root / "source"
    manifest_all = []
    for region in sorted([p for p in template_root.iterdir() if p.is_dir()]):
        region_id = region.name.upper()
        if region_filter and region_id not in region_filter:
            continue
        region_include_set = include_set
        region_file_filter = None
        if include_map is not None:
            if region_id not in include_map:
                continue
            region_include_set = set(include_map.get(region_id) or [])
        if include_files_map is not None:
            if region_id not in include_files_map:
                continue
            region_file_filter = {str(f).lower() for f in (include_files_map.get(region_id) or [])}
        agg = aggs.get(region_id)
        if not agg or not agg.get("months_present"):
            # no input data for this region; ensure out/<region> is empty
            out_region = out_dir / region_id
            out_region.mkdir(parents=True, exist_ok=True)
            if clean_out:
                for existing in out_region.glob("*.xlsx"):
                    existing.unlink()
            continue
        out_region = out_dir / region_id
        out_region.mkdir(parents=True, exist_ok=True)
        generated_names = set()
        manifest = {
            "region": region_id,
            "files": [],
            "include": sorted(region_include_set) if region_include_set else [],
            "include_files": sorted(region_file_filter) if region_file_filter else [],
        }
        template_files = [
            p for p in region.iterdir()
            if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"}
        ]
        for orig in sorted(template_files):
            name = orig.name.lower()
            core_reason = core_input_reason(orig.name)
            if core_reason:
                manifest["files"].append({
                    "file": orig.name,
                    "status": "skipped",
                    "generator": "core_input",
                    "category": "core_input",
                    "reason": core_reason,
                })
                continue
            category = classify_template(orig.name)
            if region_file_filter and orig.name.lower() not in region_file_filter:
                continue
            if not region_file_filter and region_include_set and category not in region_include_set:
                continue
            out_path = out_region / orig.name
            generated_names.add(orig.name)

            keep_vba = orig.suffix.lower() == ".xlsm"
            wb = openpyxl.load_workbook(
                orig,
                data_only=False,
                read_only=False,
                keep_links=False,
                keep_vba=keep_vba,
            )

            if "individual" in name:
                for ws in wb.worksheets:
                    fill_individual_sales(ws, agg["ind"], agg["months_present"], outlet_alias_exact, outlet_alias_name)
                manifest["files"].append({"file": orig.name, "status": "generated", "generator": "individual_sales", "category": category})
            elif "sku summary" in name:
                for ws in wb.worksheets:
                    fill_sku_summary(ws, agg["sku"], agg["months_present"], prod_alias_exact, prod_alias_name)
                manifest["files"].append({"file": orig.name, "status": "generated", "generator": "sku_summary", "category": category})
            elif "township summary" in name:
                for ws in wb.worksheets:
                    header_row, _ = find_row_with_tokens(ws, ["township"], max_scan=5)
                    if header_row:
                        fill_township_summary(ws, agg["town"], agg["months_present"], outlet_alias_exact, outlet_alias_name)
                    else:
                        fill_township_detail(ws, agg["town_prod"], ws.title, agg["months_present"], prod_alias_exact, prod_alias_name)
                manifest["files"].append({"file": orig.name, "status": "generated", "generator": "township_summary", "category": category})
            elif "van wise" in name:
                for ws in wb.worksheets:
                    fill_van_wise(ws, agg["van"], agg["months_present"], prod_alias_exact, prod_alias_name)
                manifest["files"].append({"file": orig.name, "status": "generated", "generator": "van_wise_sku", "category": category})
            else:
                # passthrough for templates we don't yet regenerate
                wb.close()
                shutil.copy2(orig, out_path)
                manifest["files"].append({"file": orig.name, "status": "passthrough", "generator": "", "category": category})
                continue

            wb.save(out_path)
            wb.close()

        manifest_path = out_region / "_regeneration_manifest.json"
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
        manifest_all.append(manifest)

        # enforce: out contains only generated files
        if clean_out:
            existing_files = [
                p for p in out_region.iterdir()
                if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"}
            ]
            for existing in existing_files:
                if existing.name not in generated_names:
                    existing.unlink()

    # write top-level manifest
    if manifest_all:
        (out_dir / "regeneration_manifest.json").write_text(
            json.dumps({"regions": manifest_all}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root-dir", default=".")
    parser.add_argument("--in-dir", default="in")
    parser.add_argument("--out-dir", default="out")
    parser.add_argument("--skip-prepare", action="store_true")
    parser.add_argument("--no-clean", action="store_true", help="Do not remove non-generated files from out/<region>")
    parser.add_argument("--include", default="", help="Comma-separated export categories to include")
    parser.add_argument("--regions", default="", help="Comma-separated regions to include")
    parser.add_argument("--include-map", default="", help="JSON mapping of region -> categories")
    parser.add_argument("--include-files-map", default="", help="JSON mapping of region -> file names")
    args = parser.parse_args()

    root = Path(args.root_dir).resolve()
    in_dir = (root / args.in_dir).resolve()
    out_dir = (root / args.out_dir).resolve()

    if not args.skip_prepare:
        prepare_in(root, in_dir)

    include = [p.strip() for p in str(args.include).split(",") if p.strip()]
    regions = [p.strip().upper() for p in str(args.regions).split(",") if p.strip()]
    include_map = json.loads(args.include_map) if args.include_map else None
    include_files_map = json.loads(args.include_files_map) if args.include_files_map else None
    regenerate_templates(
        root,
        in_dir,
        out_dir,
        clean_out=not args.no_clean,
        include=include or None,
        regions=regions or None,
        include_map=include_map,
        include_files_map=include_files_map,
    )


if __name__ == "__main__":
    main()
