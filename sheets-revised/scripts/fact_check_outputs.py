#!/usr/bin/env python3
"""Fact-check regenerated outputs against originals and rank mismatches.

Compares all sheets/rows/cols/cells between out/<REGION> and <REGION> folders.
Produces JSON + Markdown reports with rankings.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl


def norm_str(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (dt.date, dt.datetime)):
        return val.isoformat()
    s = str(val).strip()
    if s == "":
        return None
    s = re.sub(r"\s+", " ", s)
    return s.casefold()


def try_float(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, (dt.date, dt.datetime)):
        return None
    s = str(val).strip().replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None


def is_empty(val) -> bool:
    if val is None:
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    return False


def is_formula_cell(cell) -> bool:
    if cell is None:
        return False
    try:
        if cell.data_type == "f":
            return True
    except Exception:
        pass
    val = getattr(cell, "value", None)
    return isinstance(val, str) and val.startswith("=")


def norm_formula(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s.startswith("="):
        s = s[1:]
    return re.sub(r"\s+", "", s)


def best_sheet_match(target_name: str, original_names: List[str]) -> str | None:
    t = norm_str(target_name) or ""
    if not original_names:
        return None
    # exact normalized match
    for name in original_names:
        if (norm_str(name) or "") == t:
            return name
    # containment match
    for name in original_names:
        n = norm_str(name) or ""
        if t in n or n in t:
            return name
    # fallback to first if only one
    if len(original_names) == 1:
        return original_names[0]
    return None


def cell_value(cell):
    return cell.value if cell is not None else None


def load_rows(ws) -> List[Tuple]:
    rows = []
    for row in ws.iter_rows(values_only=False):
        rows.append(tuple(row))
    return rows


def trim_row(row: Tuple) -> int:
    last = 0
    for idx, cell in enumerate(row, start=1):
        if not is_empty(cell_value(cell)):
            last = idx
    return last


def compare_sheets(orig_ws, regen_ws, tol: float = 1e-6, max_cells: int | None = None) -> Dict:
    orig_rows = load_rows(orig_ws)
    regen_rows = load_rows(regen_ws)
    max_rows = max(len(orig_rows), len(regen_rows))

    stats = {
        "compared_cells": 0,
        "mismatched_cells": 0,
        "numeric_mismatches": 0,
        "string_mismatches": 0,
        "formula_mismatches": 0,
        "row_mismatches": defaultdict(int),
        "col_mismatches": defaultdict(int),
        "top_mismatches": [],  # list of dicts
    }

    def record_mismatch(r, c, v1, v2, mtype, diff=None):
        stats["mismatched_cells"] += 1
        stats["row_mismatches"][r] += 1
        stats["col_mismatches"][c] += 1
        if mtype == "numeric":
            stats["numeric_mismatches"] += 1
        elif mtype == "formula":
            stats["formula_mismatches"] += 1
        else:
            stats["string_mismatches"] += 1
        if len(stats["top_mismatches"]) < 200:
            item = {"row": r, "col": c, "type": mtype, "orig": v1, "regen": v2}
            if diff is not None:
                item["diff"] = diff
            stats["top_mismatches"].append(item)

    cell_count = 0
    for r_idx in range(1, max_rows + 1):
        o_row = orig_rows[r_idx - 1] if r_idx - 1 < len(orig_rows) else ()
        r_row = regen_rows[r_idx - 1] if r_idx - 1 < len(regen_rows) else ()
        o_len = trim_row(o_row)
        r_len = trim_row(r_row)
        max_len = max(o_len, r_len)
        if max_len == 0:
            continue
        for c_idx in range(1, max_len + 1):
            cell1 = o_row[c_idx - 1] if c_idx - 1 < len(o_row) else None
            cell2 = r_row[c_idx - 1] if c_idx - 1 < len(r_row) else None
            v1 = cell_value(cell1)
            v2 = cell_value(cell2)
            if is_empty(v1) and is_empty(v2):
                continue
            stats["compared_cells"] += 1
            cell_count += 1

            if is_formula_cell(cell1) or is_formula_cell(cell2):
                f1 = norm_formula(v1)
                f2 = norm_formula(v2)
                if f1 != f2:
                    record_mismatch(r_idx, c_idx, v1, v2, "formula")
                if max_cells and cell_count >= max_cells:
                    return stats
                continue
            f1 = try_float(v1)
            f2 = try_float(v2)
            if f1 is not None and f2 is not None:
                if abs(f1 - f2) > tol:
                    record_mismatch(r_idx, c_idx, v1, v2, "numeric", diff=f1 - f2)
            else:
                s1 = norm_str(v1)
                s2 = norm_str(v2)
                if s1 != s2:
                    record_mismatch(r_idx, c_idx, v1, v2, "string")
            if max_cells and cell_count >= max_cells:
                return stats
    return stats


def rank_dict(d: Dict[int, int]) -> List[Tuple[int, int]]:
    return sorted(d.items(), key=lambda kv: (-kv[1], kv[0]))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out-dir", default="out")
    parser.add_argument("--root-dir", default=".")
    parser.add_argument("--max-cells", type=int, default=None, help="Optional cap on compared cells per sheet")
    args = parser.parse_args()

    root = Path(args.root_dir).resolve()
    out_dir = (root / args.out_dir).resolve()

    report = {
        "files": {},
        "rankings": {
            "files": [],
            "sheets": [],
        },
    }

    for region_dir in sorted([p for p in out_dir.iterdir() if p.is_dir() and p.name.isupper()]):
        original_region = root / region_dir.name
        if not original_region.exists():
            continue
        for regen_file in sorted(region_dir.glob("*.xlsx")):
            orig_file = original_region / regen_file.name
            key = f"{region_dir.name}/{regen_file.name}"
            file_entry = {
                "region": region_dir.name,
                "regenerated": str(regen_file),
                "original": str(orig_file) if orig_file.exists() else None,
                "missing_original": not orig_file.exists(),
                "sheets": {},
                "missing_sheets": {"original_only": [], "regen_only": []},
                "totals": {"compared_cells": 0, "mismatched_cells": 0},
            }
            if not orig_file.exists():
                report["files"][key] = file_entry
                continue

            orig_wb = openpyxl.load_workbook(orig_file, data_only=False, read_only=True, keep_links=False)
            regen_wb = openpyxl.load_workbook(regen_file, data_only=False, read_only=True, keep_links=False)

            orig_names = list(orig_wb.sheetnames)
            regen_names = list(regen_wb.sheetnames)
            matched_orig = set()

            for rname in regen_names:
                oname = best_sheet_match(rname, orig_names)
                if oname is None:
                    file_entry["missing_sheets"]["regen_only"].append(rname)
                    continue
                matched_orig.add(oname)
                ows = orig_wb[oname]
                rws = regen_wb[rname]
                stats = compare_sheets(ows, rws, max_cells=args.max_cells)
                stats["row_mismatches"] = rank_dict(stats["row_mismatches"])
                stats["col_mismatches"] = rank_dict(stats["col_mismatches"])
                stats["mismatch_rate"] = (
                    stats["mismatched_cells"] / stats["compared_cells"]
                    if stats["compared_cells"] else 0.0
                )
                file_entry["sheets"][rname] = {
                    "original_sheet": oname,
                    "stats": stats,
                }
                file_entry["totals"]["compared_cells"] += stats["compared_cells"]
                file_entry["totals"]["mismatched_cells"] += stats["mismatched_cells"]

                report["rankings"]["sheets"].append({
                    "file": key,
                    "sheet": rname,
                    "original_sheet": oname,
                    "compared_cells": stats["compared_cells"],
                    "mismatched_cells": stats["mismatched_cells"],
                    "mismatch_rate": stats["mismatch_rate"],
                })

            for oname in orig_names:
                if oname not in matched_orig:
                    file_entry["missing_sheets"]["original_only"].append(oname)

            file_entry["totals"]["mismatch_rate"] = (
                file_entry["totals"]["mismatched_cells"] / file_entry["totals"]["compared_cells"]
                if file_entry["totals"]["compared_cells"] else 0.0
            )

            report["files"][key] = file_entry

            orig_wb.close()
            regen_wb.close()

    # file ranking
    for key, entry in report["files"].items():
        report["rankings"]["files"].append({
            "file": key,
            "mismatch_rate": entry["totals"].get("mismatch_rate", 0.0),
            "mismatched_cells": entry["totals"].get("mismatched_cells", 0),
            "compared_cells": entry["totals"].get("compared_cells", 0),
            "missing_original": entry.get("missing_original", False),
            "missing_sheets": entry.get("missing_sheets", {}),
        })

    report["rankings"]["files"] = sorted(
        report["rankings"]["files"],
        key=lambda x: (x["missing_original"], -x["mismatch_rate"], -x["mismatched_cells"]),
    )
    report["rankings"]["sheets"] = sorted(
        report["rankings"]["sheets"],
        key=lambda x: (-x["mismatch_rate"], -x["mismatched_cells"]),
    )

    out_json = out_dir / "fact_check_report.json"
    out_md = out_dir / "fact_check_report.md"

    out_json.write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    # Markdown summary
    lines = []
    lines.append("# Fact Check Report")
    lines.append("")
    lines.append("## File Ranking (Worst First)")
    for item in report["rankings"]["files"][:20]:
        lines.append(
            f"- {item['file']}: mismatch_rate={item['mismatch_rate']:.4f}, "
            f"mismatched={item['mismatched_cells']}, compared={item['compared_cells']}"
        )
    lines.append("")
    lines.append("## Sheet Ranking (Worst First)")
    for item in report["rankings"]["sheets"][:30]:
        lines.append(
            f"- {item['file']} :: {item['sheet']} -> mismatch_rate={item['mismatch_rate']:.4f}, "
            f"mismatched={item['mismatched_cells']}, compared={item['compared_cells']}"
        )
    lines.append("")
    lines.append("## Missing Sheets")
    for key, entry in report["files"].items():
        orig_only = entry.get("missing_sheets", {}).get("original_only", [])
        regen_only = entry.get("missing_sheets", {}).get("regen_only", [])
        if not orig_only and not regen_only:
            continue
        lines.append(f"- {key}")
        if orig_only:
            lines.append(f"- original_only: {', '.join(orig_only)}")
        if regen_only:
            lines.append(f"- regen_only: {', '.join(regen_only)}")
    lines.append("")
    lines.append("## Top Cell Mismatches (Sample)")
    for key, entry in report["files"].items():
        for sname, sdata in entry.get("sheets", {}).items():
            top = sdata["stats"]["top_mismatches"][:10]
            if not top:
                continue
            lines.append(f"- {key} :: {sname}")
            for m in top:
                if "diff" in m:
                    lines.append(
                        f"- R{m['row']}C{m['col']}: diff={m['diff']}, orig={m['orig']}, regen={m['regen']}"
                    )
                else:
                    lines.append(
                        f"- R{m['row']}C{m['col']}: orig={m['orig']}, regen={m['regen']}"
                    )

    out_md.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
