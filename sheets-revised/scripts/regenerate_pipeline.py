#!/usr/bin/env python3
"""Regenerate regional Excel reports from minimal raw sources.

Pipeline steps:
1) Prepare in/ with minimal raw source files
2) Run ETL to canonical CSVs
3) Generate report workbooks into out/<REGION>/
4) Validate regenerated outputs vs original region workbooks (coarse sum check)
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import shutil
import subprocess
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import openpyxl

# Import helpers from ETL
from etl_load_sources import infer_region_id, norm_key

RAW_SOURCES = [
    "files/7-MTL for Table and DailySales(2026 Jan to Mar).xlsx",
    "files/9-MLM Table and DailySales-2026_Feb.xlsx",
    "files/MHL 2026 Feb.xlsx",
    "files/Keng Tung - Jan New Update.xlsx",
    "files/LSH Feb PJP.xlsx",
]

REPORT_PATTERNS = {
    "individual_sales": ["individual sales", "individual sale"],
    "sku_summary": ["sku summary"],
    "township_summary": ["township summary", "township summary"],
    "van_wise_sku": ["van wise sku"],
    "op_debtor": ["op sales debtor", "op sales debtors"],
    "van_wholesale_debtor": ["van+wholesale debtor", "van-wholesale debtor"],
    "follow_up": ["follow up"],
    "sales_compare": ["sales compare"],
    "pg_daily_sales": ["pg-daily sales", "pg daily sales", "pgsales", "pg sales"],
}

VALIDATION_TOKENS = {
    "individual_sales": ["pkt", "pk", "bot", "bottle", "lit", "liter", "litre"],
    "sku_summary": ["bot", "bottle", "lit", "liter", "litre"],
    "township_summary": ["bot", "bottle", "lit", "liter", "litre"],
    "van_wise_sku": ["bot", "bottle", "lit", "liter", "litre"],
}

TOKEN_MAP = {
    "pkt": "pkt",
    "pk": "pkt",
    "pack": "pkt",
    "bot": "bot",
    "bottle": "bot",
    "lit": "lit",
    "liter": "lit",
    "litre": "lit",
}


def prepare_in(root: Path, in_dir: Path) -> List[str]:
    in_dir.mkdir(parents=True, exist_ok=True)
    copied = []
    for rel in RAW_SOURCES:
        src = root / rel
        if not src.exists():
            continue
        dest = in_dir / src.name
        if not dest.exists():
            shutil.copy2(src, dest)
            copied.append(src.name)
    return copied


def run_etl(root: Path, in_dir: Path, staging_dir: Path) -> None:
    staging_dir.mkdir(parents=True, exist_ok=True)
    cmd = [sys.executable, str(root / "scripts" / "etl_load_sources.py"), "--input", str(in_dir), "--out", str(staging_dir)]
    subprocess.run(cmd, check=True)


def read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def to_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s == "":
        return 0.0
    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def month_key(date_val, year_val, month_val) -> str:
    if date_val not in (None, ""):
        s = str(date_val)
        try:
            d = dt.datetime.fromisoformat(s)
            return f"{d.year:04d}-{d.month:02d}"
        except Exception:
            pass
        try:
            d = dt.date.fromisoformat(s)
            return f"{d.year:04d}-{d.month:02d}"
        except Exception:
            pass
    try:
        y = int(float(year_val)) if year_val not in (None, "") else None
        m = int(float(month_val)) if month_val not in (None, "") else None
        if y and m:
            return f"{y:04d}-{m:02d}"
    except Exception:
        return ""
    return ""


def write_month_pivot(
    path: Path,
    sheet_name: str,
    base_headers: List[str],
    row_labels: List[Tuple],
    months: List[str],
    measures: List[str],
    data: Dict[Tuple, Dict[str, Dict[str, float]]],
    label_fn,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header rows
    for idx, h in enumerate(base_headers, start=1):
        ws.cell(1, idx, h)
        ws.cell(2, idx, "")

    col = len(base_headers) + 1
    for month in months:
        ws.cell(1, col, month)
        if len(measures) > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(measures) - 1)
        for i, m in enumerate(measures):
            ws.cell(2, col + i, m)
        col += len(measures)

    # Data rows
    r = 3
    for row_key in row_labels:
        labels = label_fn(row_key)
        for i, val in enumerate(labels, start=1):
            ws.cell(r, i, val)
        col = len(base_headers) + 1
        for month in months:
            mdata = data.get(row_key, {}).get(month, {})
            for m in measures:
                ws.cell(r, col, round(mdata.get(m, 0.0), 6))
                col += 1
        r += 1

    wb.save(path)


def write_normalized_sheet(path: Path, sheet_name: str, headers: List[str], rows: Iterable[List]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h)
    r = 2
    for row in rows:
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
        r += 1
    wb.save(path)


def find_original(region_dir: Path, patterns: List[str]) -> Path | None:
    if not region_dir.exists():
        return None
    candidates = sorted([p for p in region_dir.iterdir() if p.suffix.lower() in {".xlsx", ".xlsm"}])
    for pat in patterns:
        for p in candidates:
            if pat in p.name.lower():
                return p
    return None


def detect_header_row(ws, tokens: List[str], max_scan: int = 20) -> Tuple[int, Dict[int, str]]:
    best_row = 0
    best_score = 0
    best_map: Dict[int, str] = {}
    max_cols = min(ws.max_column or 0, 300)
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        row_vals = [norm_key(v) for v in row[:max_cols]]
        score = sum(1 for v in row_vals if v in tokens)
        if score > best_score:
            best_score = score
            best_row = r_idx
            col_map = {}
            for idx, v in enumerate(row_vals, start=1):
                if v in tokens:
                    col_map[idx] = v
            best_map = col_map
    return best_row, best_map


def sum_original_report(path: Path, tokens: List[str]) -> Dict[str, float]:
    totals = defaultdict(float)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True, keep_links=False)
    for ws in wb.worksheets:
        header_row, col_map = detect_header_row(ws, tokens)
        if header_row == 0:
            continue
        max_rows = min(ws.max_row or 0, 5000)
        for row in ws.iter_rows(min_row=header_row + 1, max_row=max_rows, values_only=True):
            for c, token in col_map.items():
                idx = c - 1
                if idx < len(row):
                    val = row[idx]
                    if isinstance(val, (int, float)):
                        key = TOKEN_MAP.get(token, token)
                        totals[key] += float(val)
        # use first matched sheet
        break
    wb.close()
    return dict(totals)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--in-dir", default="in")
    parser.add_argument("--out-dir", default="out")
    parser.add_argument("--skip-prepare", action="store_true")
    parser.add_argument("--skip-validate", action="store_true")
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[1]
    in_dir = (root / args.in_dir).resolve()
    out_dir = (root / args.out_dir).resolve()
    staging_dir = out_dir / "staging"

    report_log = {
        "prepared_in": [],
        "regions": {},
    }

    if not args.skip_prepare:
        report_log["prepared_in"] = prepare_in(root, in_dir)

    run_etl(root, in_dir, staging_dir)

    products = read_csv(staging_dir / "products.csv")
    outlets = read_csv(staging_dir / "outlets.csv")
    sales = read_csv(staging_dir / "sales_transactions.csv")
    financials = read_csv(staging_dir / "sales_financials.csv")

    product_by_id = {p.get("product_id"): p for p in products if p.get("product_id")}
    outlet_by_id = {o.get("outlet_id"): o for o in outlets if o.get("outlet_id")}
    fin_by_txn = {f.get("txn_id"): f for f in financials if f.get("txn_id")}

    # enrich transactions
    for row in sales:
        row["region_id"] = infer_region_id(row.get("source_file", ""))
        row["month_key"] = month_key(row.get("date"), row.get("year"), row.get("month"))

    # group by region
    region_txns: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in sales:
        if row.get("region_id") and row.get("month_key"):
            region_txns[row["region_id"]].append(row)

    # Generate outputs
    for region, txns in region_txns.items():
        region_out = out_dir / region
        region_out.mkdir(parents=True, exist_ok=True)
        region_log = {"generated": {}, "skipped": []}

        # Individual Sales
        ind = defaultdict(lambda: defaultdict(lambda: {"PKT": 0.0, "BOT": 0.0, "LIT": 0.0}))
        ind_rows = set()
        months = set()
        for t in txns:
            outlet_name = t.get("outlet_name_raw") or outlet_by_id.get(t.get("outlet_id"), {}).get("outlet_name_mm", "")
            township = t.get("township_name_raw") or outlet_by_id.get(t.get("outlet_id"), {}).get("township_name", "")
            month = t.get("month_key")
            if not month:
                continue
            key = (outlet_name, township)
            ind_rows.add(key)
            months.add(month)
            ind[key][month]["PKT"] += to_float(t.get("qty_pack"))
            ind[key][month]["BOT"] += to_float(t.get("qty_bottle"))
            ind[key][month]["LIT"] += to_float(t.get("qty_liter"))

        months_sorted = sorted(months)
        ind_rows_sorted = sorted(ind_rows, key=lambda x: (x[1], x[0]))

        orig = find_original(root / region, REPORT_PATTERNS["individual_sales"])
        ind_name = orig.name if orig else f"{region}_Individual_Sales.xlsx"
        ind_path = region_out / ind_name
        write_month_pivot(
            ind_path,
            "Individual Sales",
            ["Customer Name", "Township"],
            ind_rows_sorted,
            months_sorted,
            ["PKT", "BOT", "LIT"],
            ind,
            lambda k: [k[0], k[1]],
        )
        region_log["generated"]["individual_sales"] = str(ind_path)

        # SKU Summary
        sku = defaultdict(lambda: defaultdict(lambda: {"BOT": 0.0, "LIT": 0.0}))
        sku_rows = set()
        sku_months = set()
        for t in txns:
            pid = t.get("product_id")
            if not pid:
                continue
            month = t.get("month_key")
            if not month:
                continue
            sku_rows.add(pid)
            sku_months.add(month)
            sku[pid][month]["BOT"] += to_float(t.get("qty_bottle"))
            sku[pid][month]["LIT"] += to_float(t.get("qty_liter"))

        sku_months_sorted = sorted(sku_months)
        sku_rows_sorted = sorted(sku_rows, key=lambda pid: (product_by_id.get(pid, {}).get("product_name", ""), pid))
        orig = find_original(root / region, REPORT_PATTERNS["sku_summary"])
        sku_name = orig.name if orig else f"{region}_SKU_Summary.xlsx"
        sku_path = region_out / sku_name
        write_month_pivot(
            sku_path,
            "SKU Summary",
            ["Product Name", "ML", "Packing"],
            sku_rows_sorted,
            sku_months_sorted,
            ["BOT", "LIT"],
            sku,
            lambda pid: [
                product_by_id.get(pid, {}).get("product_name", ""),
                product_by_id.get(pid, {}).get("ml", ""),
                product_by_id.get(pid, {}).get("packing", ""),
            ],
        )
        region_log["generated"]["sku_summary"] = str(sku_path)

        # Township Summary
        town = defaultdict(lambda: defaultdict(lambda: {"Bottle": 0.0, "Liter": 0.0}))
        town_rows = set()
        town_months = set()
        for t in txns:
            township = t.get("township_name_raw") or ""
            if not township:
                continue
            month = t.get("month_key")
            if not month:
                continue
            town_rows.add(township)
            town_months.add(month)
            town[township][month]["Bottle"] += to_float(t.get("qty_bottle"))
            town[township][month]["Liter"] += to_float(t.get("qty_liter"))

        town_rows_sorted = sorted(town_rows)
        town_months_sorted = sorted(town_months)
        orig = find_original(root / region, REPORT_PATTERNS["township_summary"])
        town_name = orig.name if orig else f"{region}_Township_Summary.xlsx"
        town_path = region_out / town_name
        write_month_pivot(
            town_path,
            "Township Summary",
            ["Township"],
            town_rows_sorted,
            town_months_sorted,
            ["Bottle", "Liter"],
            town,
            lambda k: [k],
        )
        region_log["generated"]["township_summary"] = str(town_path)

        # Van Wise SKU (normalized)
        van_rows = []
        for t in txns:
            route = (t.get("car_no") or "").strip()
            pid = t.get("product_id")
            month = t.get("month_key")
            if not route or not pid or not month:
                continue
            van_rows.append([
                route,
                product_by_id.get(pid, {}).get("product_name", ""),
                product_by_id.get(pid, {}).get("ml", ""),
                product_by_id.get(pid, {}).get("packing", ""),
                month,
                to_float(t.get("qty_bottle")),
                to_float(t.get("qty_liter")),
            ])
        if van_rows:
            orig = find_original(root / region, REPORT_PATTERNS["van_wise_sku"])
            van_name = orig.name if orig else f"{region}_Van_Wise_SKU.xlsx"
            van_path = region_out / van_name
            write_normalized_sheet(
                van_path,
                "Van Wise SKU",
                ["Route", "Product Name", "ML", "Packing", "Month", "Bottle", "Liter"],
                van_rows,
            )
            region_log["generated"]["van_wise_sku"] = str(van_path)
        else:
            region_log["skipped"].append("van_wise_sku")

        # Debtor summary (normalized)
        debtor = defaultdict(lambda: defaultdict(lambda: {"gross": 0.0, "outstanding": 0.0, "receivable": 0.0}))
        for t in txns:
            fin = fin_by_txn.get(t.get("txn_id"))
            if not fin:
                continue
            month = t.get("month_key")
            if not month:
                continue
            outlet_name = t.get("outlet_name_raw") or outlet_by_id.get(t.get("outlet_id"), {}).get("outlet_name_mm", "")
            if not outlet_name:
                continue
            debtor[outlet_name][month]["gross"] += to_float(fin.get("gross_amount"))
            debtor[outlet_name][month]["outstanding"] += to_float(fin.get("outstanding_balance"))
            debtor[outlet_name][month]["receivable"] += (
                to_float(fin.get("receivable_1"))
                + to_float(fin.get("receivable_2"))
                + to_float(fin.get("receivable_3"))
                + to_float(fin.get("receivable_4"))
            )

        debtor_rows = []
        for outlet, months_map in debtor.items():
            for month, vals in months_map.items():
                debtor_rows.append([
                    outlet,
                    month,
                    round(vals["gross"], 6),
                    round(vals["receivable"], 6),
                    round(vals["outstanding"], 6),
                ])
        if debtor_rows:
            orig = find_original(root / region, REPORT_PATTERNS["op_debtor"])
            debt_name = orig.name if orig else f"{region}_Debtor_Summary.xlsx"
            debt_path = region_out / debt_name
            write_normalized_sheet(
                debt_path,
                "Debtor Summary",
                ["Outlet", "Month", "Gross Amount", "Receivable Total", "Outstanding Balance"],
                debtor_rows,
            )
            region_log["generated"]["debtor_summary"] = str(debt_path)
        else:
            region_log["skipped"].append("debtor_summary")

        report_log["regions"][region] = region_log

    # Validation
    validation = {"regions": {}}
    if not args.skip_validate:
        for region, rlog in report_log["regions"].items():
            validation["regions"][region] = {}
            region_dir = root / region
            for report_type, out_path_str in rlog.get("generated", {}).items():
                if report_type not in VALIDATION_TOKENS:
                    validation["regions"][region][report_type] = {"status": "skipped", "reason": "no token map"}
                    continue
                out_path = Path(out_path_str)
                orig = find_original(region_dir, REPORT_PATTERNS.get(report_type, []))
                if not orig:
                    validation["regions"][region][report_type] = {"status": "skipped", "reason": "original not found"}
                    continue
                orig_totals = sum_original_report(orig, VALIDATION_TOKENS[report_type])
                regen_totals = sum_original_report(out_path, VALIDATION_TOKENS[report_type])
                diff = {}
                for k in set(orig_totals.keys()) | set(regen_totals.keys()):
                    diff[k] = round(regen_totals.get(k, 0.0) - orig_totals.get(k, 0.0), 6)
                validation["regions"][region][report_type] = {
                    "status": "checked",
                    "original": str(orig),
                    "regenerated": str(out_path),
                    "original_totals": orig_totals,
                    "regenerated_totals": regen_totals,
                    "diff": diff,
                }

    # Write reports
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "validation_report.json").write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    (out_dir / "pipeline_report.json").write_text(json.dumps(report_log, ensure_ascii=False, indent=2), encoding="utf-8")

    # Human readable report
    lines = []
    lines.append("# Regeneration Pipeline Report")
    lines.append("")
    lines.append("## Prepared in/")
    if report_log["prepared_in"]:
        for f in report_log["prepared_in"]:
            lines.append(f"- {f}")
    else:
        lines.append("- (no new copies; in/ already populated)")
    lines.append("")
    lines.append("## Generated Outputs")
    for region, rlog in report_log["regions"].items():
        lines.append(f"- {region}")
        for rep, p in rlog.get("generated", {}).items():
            lines.append(f"- {rep}: {p}")
        for rep in rlog.get("skipped", []):
            lines.append(f"- {rep}: SKIPPED")
    lines.append("")
    lines.append("## Validation (Coarse Sum Check)")
    for region, v in validation.get("regions", {}).items():
        lines.append(f"- {region}")
        for rep, info in v.items():
            if info.get("status") != "checked":
                lines.append(f"- {rep}: {info.get('status')} ({info.get('reason')})")
                continue
            diff = info.get("diff", {})
            diff_str = ", ".join([f"{k}={v}" for k, v in diff.items()]) if diff else "n/a"
            lines.append(f"- {rep}: diff {diff_str}")

    (out_dir / "pipeline_report.md").write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
