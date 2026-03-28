#!/usr/bin/env python3
"""ETL extract for DailySales, Table, and Outlet List sheets.

Outputs canonical CSVs under --out.
"""
import argparse
import csv
import hashlib
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl


def norm(s: Optional[str]) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_key(s: Optional[str]) -> str:
    s = norm(s).lower()
    s = re.sub(r"[;:.]", "", s)
    return s


def make_id(prefix: str, key: str) -> str:
    h = hashlib.sha1(key.encode("utf-8")).hexdigest()[:10]
    return f"{prefix}_{h}"


def infer_region_id(filename: str) -> str:
    name = filename.lower()
    if "lsh" in name or "lashio" in name:
        return "LSH"
    if "lso" in name:
        return "LSO"
    if "kt" in name or "keng tung" in name:
        return "KT"
    if "mlm" in name or "mawlamyaing" in name:
        return "MLM"
    if "mtl" in name or "meiktila" in name or "mhl" in name:
        return "MTL"
    if "ygn" in name or "yangon" in name:
        return "YGN"
    return "UNK"


PRODUCT_HEADERS = ["stock id", "particular", "ml", "qty", "column1", "sales price"]
OUTLET_HEADERS = [
    "no", "sr", "စဉ်", "ကုန်သည်", "ဆိုင်အမည်", "လိပ်စာ", "လိပ်စာ အပြည့်အစုံ", "township",
    "sales", "wholesales", "way", "car no", "car no.", "pg name", "agent", "type", "ဖုန်းနံပါတ်",
    "ဘယ်သူ့ outletလဲ", "ရင်းနှီးမှု",
]

WAY_PLAN_HEADERS = [
    "date", "day", "way", "actual way", "actual way name", "route name", "a", "b", "c", "d", "s", "total",
]

DAILY_HEADERS = {
    "year": "year",
    "month": "month",
    "date": "date",
    "today": "day_label",
    "period": "period",
    "be/whisky": "sale_class_raw",
    "column1": "sale_class_raw",
    "voucherno": "voucher_no",
    "carno": "car_no",
    "customerid": "customer_id_raw",
    "customer name": "outlet_name_raw",
    "ကုန်သည်အမည်": "outlet_name_raw",
    "township": "township_name_raw",
    "လိပ်စာ": "address_raw",
    "sales": "channel",
    "wholesales": "channel",
    "particular": "sale_type_raw",
    "stockid": "product_id_raw",
    "stockname": "stock_name_raw",
    "stock name": "stock_name_raw",
    "column2": "product_id_raw",
    "pg name": "pg_name_raw",
    "pgname": "pg_name_raw",
    "id": "customer_id_raw",
    "ml": "ml_raw",
    "ပါဝင်မှု": "participation_raw",
    "bottle": "qty_bottle",
    "salespk": "qty_pack",
    "sales pk": "qty_pack",
    "sales pkt": "qty_pack",
    "sales ctn": "qty_pack",
    "salesbot": "qty_bottle",
    "sales bottle": "qty_bottle",
    "bot": "qty_bottle",
    "sales bot": "qty_bottle",
    "sales ctns": "qty_pack",
    "liter": "qty_liter",
    "sales liter": "qty_liter",
    "litre": "qty_liter",
    "lit": "qty_liter",
    "parking": "parking_fee",
    "နှုံး": "unit_rate",
    "bonus": "unit_rate",
    "သင့်ငွေ": "gross_amount",
    "mmk": "gross_amount",
    "စာရင်းဖွင့်": "opening_balance",
    "ဈေးဟောင်းလျှော့": "old_price_discount",
    "ကော်မရှင်": "commission",
    "ဈေးလျှော့": "discount",
    "ကားခလျှော့": "transport_discount",
    "ကားခ+": "transport_add",
    "ငွေရသည့်နေ့": "payment_date_1",
    "ကြွေးရငွေ": "receivable_1",
    "ငွေရသည့်နေ့(၂)": "payment_date_2",
    "ကြွေးရငွေ2": "receivable_2",
    "ငွေရသည့်နေ့(၃)": "payment_date_3",
    "ကြွေးရငွေ3": "receivable_3",
    "ငွေရသည့်နေ့ ၄": "payment_date_4",
    "ကြွေးရငွေ4": "receivable_4",
    "အကြွေးကျန်2": "outstanding_balance",
}


def is_daily_sales_title(title_norm: str, title_key: str) -> bool:
    if title_key.startswith("dailysales") or "dailysales" in title_key:
        return True
    if "pgsales" in title_key or "pgdailysales" in title_key:
        if "summary" in title_key or "report" in title_key:
            return False
        return True
    if "pgsales" in title_norm or "pg daily sales" in title_norm:
        if "summary" in title_norm or "report" in title_norm:
            return False
        return True
    return False


def detect_header_row(ws, target_headers: List[str], max_scan: int = 15) -> Optional[int]:
    best = None
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(1, min(max_scan, max_row) + 1):
        row_vals = [norm_key(v) for v in next(ws.iter_rows(min_row=r, max_row=r, max_col=max_col, values_only=True))]
        score = sum(1 for v in row_vals if v in target_headers)
        if score and (best is None or score > best[0]):
            best = (score, r)
    return best[1] if best else None


def detect_product_blocks_from_headers(headers: List[str]) -> List[Dict[str, int]]:
    blocks = []
    for idx, v in enumerate(headers, start=1):
        if v == "stock id":
            seq = headers[idx - 1: idx - 1 + 6]
            if seq[:6] == PRODUCT_HEADERS:
                blocks.append({
                    "stock_id": idx,
                    "particular": idx + 1,
                    "ml": idx + 2,
                    "qty": idx + 3,
                    "column1": idx + 4,
                    "sales_price": idx + 5,
                })
    return blocks


def detect_outlet_block_from_headers(headers: List[str]) -> Optional[Dict[str, int]]:
    for idx, v in enumerate(headers, start=1):
        if v == "no":
            window = headers[idx - 1: idx - 1 + 10]
            if "ကုန်သည်" in window or "ဆိုင်အမည်" in window:
                mapping: Dict[str, int] = {}
                for c, hv in enumerate(headers, start=1):
                    if hv == "no":
                        mapping["outlet_code"] = c
                    elif hv in ("ကုန်သည်", "ဆိုင်အမည်"):
                        mapping["outlet_name_mm"] = c
                    elif hv in ("လိပ်စာ", "လိပ်စာ အပြည့်အစုံ"):
                        mapping["address_full"] = c
                    elif hv == "township":
                        mapping["township_name"] = c
                    elif hv in ("sales", "wholesales", "type"):
                        mapping["outlet_type"] = c
                    elif hv in ("car no", "car no."):
                        mapping["way_code"] = c
                    elif hv == "pg name":
                        mapping["responsible_person"] = c
                if mapping:
                    return mapping
    return None


def parse_table_sheet(path: Path, ws, region_id: str,
                      products: Dict[str, dict], outlets: Dict[str, dict],
                      townships: Dict[str, dict], routes: Dict[str, dict], warnings: List[dict]):
    header_row = 1
    header_vals = [norm_key(v) for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    prod_blocks = detect_product_blocks_from_headers(header_vals)
    outlet_block = detect_outlet_block_from_headers(header_vals)

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        # product blocks
        for block in prod_blocks:
            def get(col_idx: int):
                return row[col_idx - 1] if col_idx - 1 < len(row) else None
            stock_id = norm(get(block["stock_id"]))
            particular = norm(get(block["particular"]))
            ml = get(block["ml"])
            packing = norm(get(block["qty"]))
            category = norm(get(block["column1"]))
            sales_price = get(block["sales_price"])
            if not stock_id and not particular:
                continue
            key = stock_id or f"{particular}|{ml}|{packing}"
            pid = stock_id or make_id("prod", norm_key(key))
            if pid not in products:
                products[pid] = {
                    "product_id": pid,
                    "product_name": particular,
                    "ml": ml,
                    "packing": packing,
                    "sales_price": sales_price,
                    "unit_type": "",
                    "category": category,
                    "source_file": path.name,
                    "source_sheet": ws.title,
                    "source_row": row_idx,
                }

        # outlet block
        if outlet_block:
            def get(col_idx: int):
                return row[col_idx - 1] if col_idx - 1 < len(row) else None
            name = norm(get(outlet_block.get("outlet_name_mm"))) if outlet_block.get("outlet_name_mm") else ""
            if not name:
                continue
            outlet_code = norm(get(outlet_block.get("outlet_code"))) if outlet_block.get("outlet_code") else ""
            address = norm(get(outlet_block.get("address_full"))) if outlet_block.get("address_full") else ""
            township = norm(get(outlet_block.get("township_name"))) if outlet_block.get("township_name") else ""
            outlet_type = norm(get(outlet_block.get("outlet_type"))) if outlet_block.get("outlet_type") else ""
            way_code = norm(get(outlet_block.get("way_code"))) if outlet_block.get("way_code") else ""
            responsible = norm(get(outlet_block.get("responsible_person"))) if outlet_block.get("responsible_person") else ""

            if township:
                tid = make_id("town", norm_key(f"{region_id}|{township}"))
                if tid not in townships:
                    townships[tid] = {
                        "township_id": tid,
                        "township_name": township,
                        "township_name_en": "",
                        "region_id": region_id,
                        "source_file": path.name,
                    }
            key = norm_key(f"{name}|{township}|{outlet_code}")
            oid = make_id("out", key)
            if oid not in outlets:
                outlets[oid] = {
                    "outlet_id": oid,
                    "outlet_code": outlet_code,
                    "outlet_name_mm": name,
                    "outlet_name_en": "",
                    "outlet_type": outlet_type,
                    "address_full": address,
                    "township_name": township,
                    "way_code": way_code,
                    "contact_phone": "",
                    "agent_name": "",
                    "responsible_person": responsible,
                    "notes": "",
                    "source_file": path.name,
                    "source_sheet": ws.title,
                    "source_row": row_idx,
                }
            if way_code:
                rid = make_id("route", norm_key(f"{region_id}|{way_code}"))
                if rid not in routes:
                    routes[rid] = {
                        "route_id": rid,
                        "region_id": region_id,
                        "van_id": "",
                        "way_code": way_code,
                        "route_name": "",
                        "source_file": path.name,
                    }


def parse_outlet_list_sheet(path: Path, ws, region_id: str,
                            outlets: Dict[str, dict], townships: Dict[str, dict], routes: Dict[str, dict]):
    header_row = detect_header_row(ws, OUTLET_HEADERS)
    if not header_row:
        return
    header_vals = [norm_key(v) for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]

    start_cols = [i for i, hv in enumerate(header_vals, start=1) if hv in ("sr", "no", "စဉ်")]
    blocks: List[Tuple[int, int]] = []
    if len(start_cols) >= 2:
        block_width = start_cols[1] - start_cols[0]
        if block_width > 1:
            for s in start_cols:
                e = s + block_width - 1
                if e <= len(header_vals):
                    blocks.append((s, e))
    if not blocks:
        blocks = [(1, len(header_vals))]

    def build_header_map(start: int, end: int) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for c in range(start, end + 1):
            hv = header_vals[c - 1]
            if hv in ("sr", "no", "စဉ်"):
                mapping["outlet_code"] = c
            elif hv in ("ကုန်သည်", "ဆိုင်အမည်"):
                mapping["outlet_name_mm"] = c
            elif hv in ("type", "type "):
                mapping["outlet_type"] = c
            elif hv in ("လိပ်စာ", "လိပ်စာ အပြည့်အစုံ"):
                mapping["address_full"] = c
            elif hv == "township":
                mapping["township_name"] = c
            elif hv == "way":
                mapping["way_code"] = c
            elif hv == "ဖုန်းနံပါတ်":
                mapping["contact_phone"] = c
            elif hv == "agent":
                mapping["agent_name"] = c
            elif hv == "ဘယ်သူ့ outletလဲ":
                mapping["responsible_person"] = c
            elif hv == "ရင်းနှီးမှု":
                mapping["notes"] = c
        return mapping

    header_maps = [build_header_map(s, e) for s, e in blocks]

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        for header_map in header_maps:
            def get(col_idx: int):
                return row[col_idx - 1] if col_idx - 1 < len(row) else None
            name = norm(get(header_map.get("outlet_name_mm"))) if header_map.get("outlet_name_mm") else ""
            if not name:
                continue
            outlet_code = norm(get(header_map.get("outlet_code"))) if header_map.get("outlet_code") else ""
            address = norm(get(header_map.get("address_full"))) if header_map.get("address_full") else ""
            township = norm(get(header_map.get("township_name"))) if header_map.get("township_name") else ""
            outlet_type = norm(get(header_map.get("outlet_type"))) if header_map.get("outlet_type") else ""
            way_code = norm(get(header_map.get("way_code"))) if header_map.get("way_code") else ""
            contact_phone = norm(get(header_map.get("contact_phone"))) if header_map.get("contact_phone") else ""
            agent_name = norm(get(header_map.get("agent_name"))) if header_map.get("agent_name") else ""
            responsible = norm(get(header_map.get("responsible_person"))) if header_map.get("responsible_person") else ""
            notes = norm(get(header_map.get("notes"))) if header_map.get("notes") else ""

            if township:
                tid = make_id("town", norm_key(f"{region_id}|{township}"))
                if tid not in townships:
                    townships[tid] = {
                        "township_id": tid,
                        "township_name": township,
                        "township_name_en": "",
                        "region_id": region_id,
                        "source_file": path.name,
                    }
            key = norm_key(f"{name}|{township}|{contact_phone}")
            oid = make_id("out", key)
            if oid not in outlets:
                outlets[oid] = {
                    "outlet_id": oid,
                    "outlet_code": outlet_code,
                    "outlet_name_mm": name,
                    "outlet_name_en": "",
                    "outlet_type": outlet_type,
                    "address_full": address,
                    "township_name": township,
                    "way_code": way_code,
                    "contact_phone": contact_phone,
                    "agent_name": agent_name,
                    "responsible_person": responsible,
                    "notes": notes,
                    "source_file": path.name,
                    "source_sheet": ws.title,
                    "source_row": row_idx,
                }
            if way_code:
                rid = make_id("route", norm_key(f"{region_id}|{way_code}"))
                if rid not in routes:
                    routes[rid] = {
                        "route_id": rid,
                        "region_id": region_id,
                        "van_id": "",
                        "way_code": way_code,
                        "route_name": "",
                        "source_file": path.name,
                    }


def parse_daily_sales_sheet(path: Path, ws, region_id: str,
                            products: Dict[str, dict], outlets: Dict[str, dict], townships: Dict[str, dict],
                            sales_rows: List[dict], financial_rows: List[dict]):
    target_headers = [norm_key(h) for h in DAILY_HEADERS.keys()]
    header_row = detect_header_row(ws, target_headers, max_scan=30)
    if not header_row:
        return
    header_vals = [norm_key(v) for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    header_to_col: Dict[str, int] = {}
    for idx, hv in enumerate(header_vals, start=1):
        if hv and hv not in header_to_col:
            header_to_col[hv] = idx
    match_count = sum(1 for hv in header_to_col if hv in DAILY_HEADERS)
    if match_count < 5:
        return
    header_keys = set(DAILY_HEADERS.keys())

    def row_header_match(row_vals: List[str]) -> int:
        return sum(1 for v in row_vals if v in header_keys)

    def get_col(hv: str, default_idx: int):
        col = header_to_col.get(hv)
        if col:
            return row[col - 1] if col - 1 < len(row) else None
        return row[default_idx] if len(row) > default_idx else None

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        row_vals = [norm_key(v) for v in row]
        if row_header_match(row_vals) >= 5:
            header_to_col = {}
            for idx, hv in enumerate(row_vals, start=1):
                if hv and hv not in header_to_col:
                    header_to_col[hv] = idx
            continue
        year = get_col("year", 0)
        month = get_col("month", 1)
        date_val = get_col("date", 2)
        if year in (None, "") and month in (None, "") and date_val in (None, ""):
            continue

        row_data = {}
        fin_data = {}
        for hv, c in header_to_col.items():
            if c - 1 >= len(row):
                continue
            value = row[c - 1]
            key = DAILY_HEADERS.get(hv)
            if not key:
                continue
            if key in (
                "unit_rate", "gross_amount", "opening_balance", "old_price_discount", "commission", "discount",
                "transport_discount", "transport_add", "payment_date_1", "receivable_1", "payment_date_2",
                "receivable_2", "payment_date_3", "receivable_3", "payment_date_4", "receivable_4",
                "outstanding_balance",
            ):
                fin_data[key] = value
            else:
                row_data[key] = value

        # build date
        date_final = None
        if hasattr(date_val, "date"):
            date_final = date_val.date()
        else:
            try:
                y = int(year) if year not in (None, "") else None
                m = int(month) if month not in (None, "") else None
                d = int(date_val) if date_val not in (None, "") else None
                if y and m and d:
                    date_final = f"{y:04d}-{m:02d}-{d:02d}"
            except Exception:
                date_final = None

        outlet_name = norm(row_data.get("outlet_name_raw"))
        township = norm(row_data.get("township_name_raw"))
        if not township and row_data.get("pg_name_raw") and row_data.get("channel"):
            township = norm(row_data.get("channel"))
            row_data["township_name_raw"] = township
            row_data["channel"] = ""
        if not outlet_name and row_data.get("customer_id_raw"):
            outlet_name = norm(row_data.get("customer_id_raw"))
        stock_id_raw = norm(row_data.get("product_id_raw"))
        stock_name_raw = norm(row_data.get("stock_name_raw"))
        ml_raw = row_data.get("ml_raw")

        if not (outlet_name or row_data.get("customer_id_raw")):
            continue
        if not (stock_id_raw or stock_name_raw):
            continue

        if township:
            tid = make_id("town", norm_key(f"{region_id}|{township}"))
            if tid not in townships:
                townships[tid] = {
                    "township_id": tid,
                    "township_name": township,
                    "township_name_en": "",
                    "region_id": region_id,
                    "source_file": path.name,
                }

        if stock_id_raw and stock_id_raw in products:
            product_id = stock_id_raw
        elif stock_id_raw:
            product_id = stock_id_raw
            if product_id not in products:
                products[product_id] = {
                    "product_id": product_id,
                    "product_name": stock_name_raw,
                    "ml": ml_raw,
                    "packing": "",
                    "sales_price": "",
                    "unit_type": "",
                    "category": "",
                    "source_file": path.name,
                    "source_sheet": ws.title,
                    "source_row": row_idx,
                }
        else:
            if stock_name_raw:
                product_id = make_id("prod", norm_key(f"{stock_name_raw}|{ml_raw}"))
                if product_id not in products:
                    products[product_id] = {
                        "product_id": product_id,
                        "product_name": stock_name_raw,
                        "ml": ml_raw,
                        "packing": "",
                        "sales_price": "",
                        "unit_type": "",
                        "category": "",
                        "source_file": path.name,
                        "source_sheet": ws.title,
                        "source_row": row_idx,
                    }
            else:
                product_id = ""

        outlet_id = ""
        if outlet_name:
            key = norm_key(f"{outlet_name}|{township}")
            outlet_id = make_id("out", key)
            if outlet_id not in outlets:
                outlets[outlet_id] = {
                    "outlet_id": outlet_id,
                    "outlet_code": "",
                    "outlet_name_mm": outlet_name,
                    "outlet_name_en": "",
                    "outlet_type": "",
                    "address_full": norm(row_data.get("address_raw")),
                    "township_name": township,
                    "way_code": "",
                    "contact_phone": "",
                    "agent_name": "",
                    "responsible_person": "",
                    "notes": "",
                    "source_file": path.name,
                    "source_sheet": ws.title,
                    "source_row": row_idx,
                }
            # if car_no missing, backfill from outlet way_code when available
            if not row_data.get("car_no") and outlets.get(outlet_id, {}).get("way_code"):
                row_data["car_no"] = outlets[outlet_id].get("way_code")
        if not row_data.get("car_no") and row_data.get("pg_name_raw"):
            row_data["car_no"] = norm(row_data.get("pg_name_raw"))

        txn_key = f"{path.name}|{ws.title}|{row_idx}"
        txn_id = make_id("txn", txn_key)
        sales_rows.append({
            "txn_id": txn_id,
            "date": date_final,
            "year": row_data.get("year"),
            "month": row_data.get("month"),
            "day": row_data.get("date"),
            "day_label": row_data.get("day_label"),
            "period": row_data.get("period"),
            "outlet_id": outlet_id,
            "customer_id_raw": row_data.get("customer_id_raw"),
            "outlet_name_raw": outlet_name,
            "township_name_raw": township,
            "product_id": product_id,
            "stock_id_raw": stock_id_raw,
            "stock_name_raw": stock_name_raw,
            "ml_raw": ml_raw,
            "packing_raw": row_data.get("packing_raw"),
            "channel": row_data.get("channel"),
            "voucher_no": row_data.get("voucher_no"),
            "car_no": row_data.get("car_no"),
            "qty_pack": row_data.get("qty_pack"),
            "qty_bottle": row_data.get("qty_bottle"),
            "qty_liter": row_data.get("qty_liter"),
            "sale_type_raw": row_data.get("sale_type_raw"),
            "sale_class_raw": row_data.get("sale_class_raw"),
            "participation_raw": row_data.get("participation_raw"),
            "parking_fee": row_data.get("parking_fee"),
            "source_file": path.name,
            "source_sheet": ws.title,
            "source_row": row_idx,
        })

        if fin_data:
            fin_data["txn_id"] = txn_id
            financial_rows.append(fin_data)


def parse_way_plan_sheet(path: Path, ws, region_id: str,
                         routes: Dict[str, dict], pjp_rows: List[dict]):
    target_headers = [norm_key(h) for h in WAY_PLAN_HEADERS]
    header_row = detect_header_row(ws, target_headers, max_scan=20)
    if not header_row:
        return
    header_vals = [norm_key(v) for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    header_to_col: Dict[str, int] = {}
    for idx, hv in enumerate(header_vals, start=1):
        if hv and hv not in header_to_col:
            header_to_col[hv] = idx

    def get_col(row, hv: str, default_idx: int | None = None):
        col = header_to_col.get(hv)
        if col:
            return row[col - 1] if col - 1 < len(row) else None
        if default_idx is not None and len(row) > default_idx:
            return row[default_idx]
        return None

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        date_val = get_col(row, "date", 0)
        way = norm(get_col(row, "way", 2))
        route_name = norm(get_col(row, "actual way name", 3) or get_col(row, "route name", 3))
        if not date_val and not way and not route_name:
            continue
        if not way and not route_name:
            continue

        # normalize date
        if hasattr(date_val, "date"):
            date_final = date_val.date().isoformat()
        else:
            date_final = str(date_val) if date_val not in (None, "") else ""

        route_key = norm_key(f"{region_id}|{way or route_name}")
        route_id = make_id("route", route_key)
        if route_id not in routes:
            routes[route_id] = {
                "route_id": route_id,
                "region_id": region_id,
                "van_id": "",
                "way_code": way,
                "route_name": route_name,
                "source_file": path.name,
            }

        a = get_col(row, "a")
        b = get_col(row, "b")
        c = get_col(row, "c")
        d = get_col(row, "d")
        s = get_col(row, "s")
        total = get_col(row, "total")

        plan_id = make_id("plan", norm_key(f"{region_id}|{route_id}|{date_final}|{row_idx}"))
        pjp_rows.append({
            "plan_id": plan_id,
            "date": date_final,
            "route_id": route_id,
            "planned_a": a,
            "planned_b": b,
            "planned_c": c,
            "planned_d": d,
            "planned_s": s,
            "total_planned": total,
            "source_file": path.name,
            "source_sheet": ws.title,
            "source_row": row_idx,
        })


def parse_pjp_outlets_sheet(path: Path, ws, region_id: str,
                            outlets: Dict[str, dict], routes: Dict[str, dict],
                            route_outlets: List[dict]):
    title_key = norm_key(ws.title)
    if "pjp" not in title_key or "outlet" not in title_key:
        return
    # attempt to infer route name from top rows
    route_name = ""
    for r in range(1, 8):
        for c in range(1, 8):
            cell_val = norm(ws.cell(row=r, column=c).value)
            if not cell_val:
                continue
            if "route" in cell_val.lower() or "way" in cell_val.lower():
                if ":" in cell_val:
                    route_name = cell_val.split(":", 1)[-1].strip()
                else:
                    next_val = norm(ws.cell(row=r, column=c + 1).value)
                    route_name = next_val or cell_val
                break
        if route_name:
            break

    header_row = detect_header_row(ws, OUTLET_HEADERS, max_scan=30)
    if not header_row:
        return
    header_vals = [norm_key(v) for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    mapping: Dict[str, int] = {}
    for c, hv in enumerate(header_vals, start=1):
        if hv in ("sr", "no", "စဉ်"):
            mapping["outlet_code"] = c
        elif hv in ("ကုန်သည်", "ဆိုင်အမည်"):
            mapping["outlet_name_mm"] = c
        elif hv in ("type", "type "):
            mapping["outlet_type"] = c
        elif hv in ("လိပ်စာ", "လိပ်စာ အပြည့်အစုံ"):
            mapping["address_full"] = c
        elif hv == "township":
            mapping["township_name"] = c
        elif hv in ("ဖုန်းနံပါတ်",):
            mapping["contact_phone"] = c

    route_key = norm_key(f"{region_id}|{route_name}")
    route_id = make_id("route", route_key) if route_name else ""
    if route_id and route_id not in routes:
        routes[route_id] = {
            "route_id": route_id,
            "region_id": region_id,
            "van_id": "",
            "way_code": "",
            "route_name": route_name,
            "source_file": path.name,
        }

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        def get(col_idx: int):
            return row[col_idx - 1] if col_idx - 1 < len(row) else None
        name = norm(get(mapping.get("outlet_name_mm", 0))) if mapping.get("outlet_name_mm") else ""
        if not name:
            continue
        township = norm(get(mapping.get("township_name", 0))) if mapping.get("township_name") else ""
        outlet_type = norm(get(mapping.get("outlet_type", 0))) if mapping.get("outlet_type") else ""
        contact_phone = norm(get(mapping.get("contact_phone", 0))) if mapping.get("contact_phone") else ""
        address = norm(get(mapping.get("address_full", 0))) if mapping.get("address_full") else ""

        # create outlet if missing
        key = norm_key(f"{name}|{township}|{contact_phone}")
        oid = make_id("out", key)
        if oid not in outlets:
            outlets[oid] = {
                "outlet_id": oid,
                "outlet_code": "",
                "outlet_name_mm": name,
                "outlet_name_en": "",
                "outlet_type": outlet_type,
                "address_full": address,
                "township_name": township,
                "way_code": "",
                "contact_phone": contact_phone,
                "agent_name": "",
                "responsible_person": "",
                "notes": "",
                "source_file": path.name,
                "source_sheet": ws.title,
                "source_row": row_idx,
            }

        if route_id:
            route_outlets.append({
                "route_id": route_id,
                "outlet_id": oid,
                "category": "",
                "start_date": "",
                "end_date": "",
                "source_file": path.name,
                "source_sheet": ws.title,
                "source_row": row_idx,
            })


def write_csv(path: Path, rows: List[dict], fieldnames: List[str]):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Input directory with .xlsx files")
    parser.add_argument("--out", required=True, help="Output directory for CSVs")
    args = parser.parse_args()

    in_dir = Path(args.input)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    products: Dict[str, dict] = {}
    outlets: Dict[str, dict] = {}
    townships: Dict[str, dict] = {}
    routes: Dict[str, dict] = {}
    sales_rows: List[dict] = []
    financial_rows: List[dict] = []
    pjp_rows: List[dict] = []
    route_outlets: List[dict] = []

    files = sorted([p for p in in_dir.iterdir() if p.suffix.lower() in {".xlsx", ".xlsm"}])
    for path in files:
        region_id = infer_region_id(path.name)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True, keep_links=False)
        for ws in wb.worksheets:
            title_norm = norm_key(ws.title)
            title_key = re.sub(r"[^a-z0-9]+", "", title_norm)
            handled = False
            if title_key.startswith("table"):
                parse_table_sheet(path, ws, region_id, products, outlets, townships, routes, [])
                handled = True
            elif is_daily_sales_title(title_norm, title_key):
                parse_daily_sales_sheet(path, ws, region_id, products, outlets, townships, sales_rows, financial_rows)
                handled = True
            elif "outletlist" in title_key:
                parse_outlet_list_sheet(path, ws, region_id, outlets, townships, routes)
                handled = True
            elif "wayplan" in title_key or "wayplay" in title_key or "wayplan" in title_norm or "wayplay" in title_norm:
                parse_way_plan_sheet(path, ws, region_id, routes, pjp_rows)
                handled = True
            elif "pjpoutlets" in title_key or ("pjp" in title_key and "outlet" in title_key):
                parse_pjp_outlets_sheet(path, ws, region_id, outlets, routes, route_outlets)
                handled = True
            if not handled:
                # Fallback: attempt daily sales parsing based on headers (handles non-standard sheet names)
                parse_daily_sales_sheet(path, ws, region_id, products, outlets, townships, sales_rows, financial_rows)
        wb.close()

    write_csv(out_dir / "products.csv", list(products.values()), [
        "product_id", "product_name", "ml", "packing", "sales_price", "unit_type", "category",
        "source_file", "source_sheet", "source_row",
    ])
    write_csv(out_dir / "outlets.csv", list(outlets.values()), [
        "outlet_id", "outlet_code", "outlet_name_mm", "outlet_name_en", "outlet_type", "address_full",
        "township_name", "way_code", "contact_phone", "agent_name", "responsible_person", "notes",
        "source_file", "source_sheet", "source_row",
    ])
    write_csv(out_dir / "townships.csv", list(townships.values()), [
        "township_id", "township_name", "township_name_en", "region_id", "source_file",
    ])
    write_csv(out_dir / "routes.csv", list(routes.values()), [
        "route_id", "region_id", "van_id", "way_code", "route_name", "source_file",
    ])
    write_csv(out_dir / "sales_transactions.csv", sales_rows, [
        "txn_id", "date", "year", "month", "day", "day_label", "period",
        "outlet_id", "customer_id_raw", "outlet_name_raw", "township_name_raw",
        "product_id", "stock_id_raw", "stock_name_raw", "ml_raw", "packing_raw",
        "channel", "voucher_no", "car_no",
        "qty_pack", "qty_bottle", "qty_liter",
        "sale_type_raw", "sale_class_raw", "participation_raw", "parking_fee",
        "source_file", "source_sheet", "source_row",
    ])
    write_csv(out_dir / "sales_financials.csv", financial_rows, [
        "txn_id", "unit_rate", "gross_amount", "opening_balance", "old_price_discount", "commission",
        "discount", "transport_discount", "transport_add", "payment_date_1", "receivable_1",
        "payment_date_2", "receivable_2", "payment_date_3", "receivable_3", "payment_date_4",
        "receivable_4", "outstanding_balance",
    ])

    write_csv(out_dir / "pjp_plans.csv", pjp_rows, [
        "plan_id", "date", "route_id", "planned_a", "planned_b", "planned_c",
        "planned_d", "planned_s", "total_planned", "source_file", "source_sheet", "source_row",
    ])

    write_csv(out_dir / "route_outlets.csv", route_outlets, [
        "route_id", "outlet_id", "category", "start_date", "end_date", "source_file", "source_sheet", "source_row",
    ])


if __name__ == "__main__":
    main()
