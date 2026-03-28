from __future__ import annotations

import datetime as dt
import uuid
import json
import io
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import FileResponse, JSONResponse, Response
import openpyxl

from app.core.config import EXPORT_ROOT
from app.core.db import get_conn, row_to_dict
from app.services.excel_export import build_export_zip, list_regions, prepare_export_dir, run_excel_regeneration


EXPORT_CATEGORIES = {
    "individual_sales",
    "sku_summary",
    "township_summary",
    "van_wise_sku",
    "sales_compare",
    "follow_up_sales",
    "debtors",
    "other",
}

router = APIRouter(prefix="/api")


def list_export_runs(limit: int = 10) -> List[Dict[str, Any]]:
    runs: List[Dict[str, Any]] = []
    if not EXPORT_ROOT.exists():
        return runs
    candidates = [p for p in EXPORT_ROOT.iterdir() if p.is_dir() and p.name.startswith("export_")]
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    for p in candidates[:limit]:
        runs.append({
            "id": p.name,
            "path": str(p),
            "updated_at": dt.datetime.utcfromtimestamp(p.stat().st_mtime).isoformat(),
        })
    return runs


def _cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (dt.date, dt.datetime)):
        return val.isoformat()
    return str(val).strip()


def _preview_workbook_bytes(payload: bytes, name: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        sheets = []
        for ws in wb.worksheets:
            header = []
            header_row_index = None
            for idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
                if not row:
                    continue
                if any(cell not in (None, "") for cell in row):
                    header = [_cell_to_str(cell) for cell in row]
                    header_row_index = idx
                    break
            while header and header[-1] == "":
                header.pop()
            sample_rows = []
            max_cols = min(max(len(header), 6), 12)
            start_row = header_row_index + 1 if header_row_index else 1
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + 4, values_only=True):
                if not row:
                    continue
                sample_rows.append([_cell_to_str(cell) for cell in row[:max_cols]])
            sheets.append({
                "name": ws.title,
                "rows": ws.max_row or 0,
                "cols": ws.max_column or 0,
                "header": header,
                "sample_rows": sample_rows,
            })
    finally:
        wb.close()
    return {"file": name, "sheets": sheets}


def _preview_workbook(path: Path) -> Dict[str, Any]:
    return _preview_workbook_bytes(path.read_bytes(), path.name)


def _read_sheet_data(path: Path, sheet: str, offset: int = 0, limit: int = 50, q: str | None = None) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet not found")
        ws = wb[sheet]
        header = []
        header_row_index = None
        for idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
            if not row:
                continue
            if any(cell not in (None, "") for cell in row):
                header = [_cell_to_str(cell) for cell in row]
                header_row_index = idx
                break
        while header and header[-1] == "":
            header.pop()
        start_row = header_row_index + 1 if header_row_index else 1
        rows = []
        total = 0
        query = (q or "").lower().strip()
        if not query:
            total = max((ws.max_row or 0) - start_row + 1, 0)
            for idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=0):
                if idx < offset:
                    continue
                if len(rows) >= limit:
                    break
                rows.append([_cell_to_str(cell) for cell in row])
        else:
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not row:
                    continue
                row_vals = [_cell_to_str(cell) for cell in row]
                hay = " ".join(row_vals).lower()
                if query not in hay:
                    continue
                if total >= offset and len(rows) < limit:
                    rows.append(row_vals)
                total += 1
        return {
            "sheet": sheet,
            "header": header,
            "rows": rows,
            "total": total,
            "offset": offset,
            "limit": limit,
        }
    finally:
        wb.close()


def load_regeneration_manifest(run_dir: Path) -> Optional[Dict[str, Any]]:
    top_manifest = run_dir / "regeneration_manifest.json"
    if top_manifest.exists():
        try:
            return json.loads(top_manifest.read_text(encoding="utf-8"))
        except Exception:
            return None
    regions = []
    for region_dir in sorted([p for p in run_dir.iterdir() if p.is_dir() and p.name.isupper()]):
        manifest_path = region_dir / "_regeneration_manifest.json"
        if not manifest_path.exists():
            continue
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        regions.append(manifest)
    if regions:
        return {"regions": regions}
    return None


def ensure_export_table(conn) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS export_jobs (
          export_id TEXT PRIMARY KEY,
          requested_at TEXT NOT NULL,
          completed_at TEXT,
          region TEXT,
          status TEXT NOT NULL,
          file_path TEXT,
          error_message TEXT,
          requested_by TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS ix_export_jobs_requested_at
          ON export_jobs (requested_at)
        """
    )


@router.get("/regions")
def regions():
    return JSONResponse({"regions": list_regions()})


@router.get("/reports/summary")
def report_summary(start: str, end: str):
    conn = get_conn()
    try:
        row = conn.execute(
            """
            SELECT
              COALESCE(SUM(qty_pack), 0) AS total_pack,
              COALESCE(SUM(qty_bottle), 0) AS total_bottle,
              COALESCE(SUM(qty_liter), 0) AS total_liter,
              COUNT(DISTINCT outlet_id) AS active_outlets
            FROM sales_transactions
            WHERE date BETWEEN ? AND ?
            """,
            (start, end),
        ).fetchone()
        return JSONResponse({"start": start, "end": end, **row_to_dict(row)})
    finally:
        conn.close()


@router.get("/reports/trends")
def report_trends(start: str, end: str):
    conn = get_conn()
    try:
        rows = conn.execute(
            """
            SELECT date,
                   COALESCE(SUM(qty_pack), 0) AS total_pack,
                   COALESCE(SUM(qty_liter), 0) AS total_liter,
                   COUNT(DISTINCT outlet_id) AS active_outlets
            FROM sales_transactions
            WHERE date BETWEEN ? AND ?
            GROUP BY date
            ORDER BY date
            """,
            (start, end),
        ).fetchall()
        return JSONResponse({"start": start, "end": end, "rows": [row_to_dict(r) for r in rows]})
    finally:
        conn.close()


@router.get("/reports/sales_export")
def sales_export(start: str, end: str):
    conn = get_conn()
    try:
        rows = conn.execute(
            """
            SELECT s.date, s.outlet_id, o.outlet_name_mm, o.outlet_name_en,
                   s.product_id, p.product_name, s.qty_pack, s.qty_bottle, s.qty_liter,
                   s.channel, s.voucher_no, s.car_no, s.route_id
            FROM sales_transactions s
            LEFT JOIN outlets o ON s.outlet_id = o.outlet_id
            LEFT JOIN products p ON s.product_id = p.product_id
            WHERE s.date BETWEEN ? AND ?
            ORDER BY s.date, s.outlet_id
            """,
            (start, end),
        ).fetchall()

        header = [
            "date", "outlet_id", "outlet_name_mm", "outlet_name_en",
            "product_id", "product_name", "qty_pack", "qty_bottle", "qty_liter",
            "channel", "voucher_no", "car_no", "route_id",
        ]
        lines = [",".join(header)]
        for r in rows:
            vals = [row_to_dict(r).get(h, "") for h in header]
            out = []
            for v in vals:
                v = "" if v is None else str(v)
                if "," in v or "\"" in v or "\n" in v:
                    v = "\"" + v.replace("\"", "\"\"") + "\""
                out.append(v)
            lines.append(",".join(out))
        csv_data = "\n".join(lines)
        return Response(
            content=csv_data,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=\"sales_{start}_to_{end}.csv\""},
        )
    finally:
        conn.close()


@router.get("/quality/summary")
def quality_summary(start: str | None = None, end: str | None = None):
    conn = get_conn()
    try:
        where = ""
        params: List[Any] = []
        if start and end:
            where = "WHERE date BETWEEN ? AND ?"
            params.extend([start, end])
        row = conn.execute(
            f"""
            SELECT
              COUNT(*) AS total_rows,
              SUM(CASE WHEN outlet_id IS NULL OR outlet_id = '' THEN 1 ELSE 0 END) AS missing_outlet,
              SUM(CASE WHEN product_id IS NULL OR product_id = '' THEN 1 ELSE 0 END) AS missing_product,
              SUM(CASE WHEN date IS NULL THEN 1 ELSE 0 END) AS missing_date,
              SUM(CASE WHEN qty_liter IS NULL THEN 1 ELSE 0 END) AS missing_liter,
              SUM(CASE WHEN route_id IS NULL OR route_id = '' THEN 1 ELSE 0 END) AS missing_route
            FROM sales_transactions
            {where}
            """,
            params,
        ).fetchone()
        return JSONResponse({"start": start, "end": end, **row_to_dict(row)})
    finally:
        conn.close()


@router.post("/reports/export_excel")
async def export_excel(request: Request):
    payload = await request.json()
    region = payload.get("region") if isinstance(payload, dict) else None
    regions = payload.get("regions") if isinstance(payload, dict) else None
    include = payload.get("include") if isinstance(payload, dict) else None
    selections = payload.get("selections") if isinstance(payload, dict) else None
    if include is not None and isinstance(include, str):
        include = [v.strip() for v in include.split(",") if v.strip()]
    if include is not None and not isinstance(include, list):
        raise HTTPException(status_code=400, detail="include must be a list or comma-separated string")
    if include:
        include = [str(v) for v in include]
        invalid = [v for v in include if v not in EXPORT_CATEGORIES]
        if invalid:
            raise HTTPException(status_code=400, detail=f"Unknown export categories: {', '.join(invalid)}")

    include_map: Dict[str, List[str]] | None = None
    include_files_map: Dict[str, List[str]] | None = None
    if selections:
        if not isinstance(selections, list):
            raise HTTPException(status_code=400, detail="selections must be a list")
        include_map = {}
        include_files_map = {}
        for item in selections:
            if not isinstance(item, dict):
                continue
            reg = str(item.get("region") or "").upper()
            cats = item.get("categories")
            files = item.get("files")
            if not reg:
                continue
            if not cats:
                cats = None
            if isinstance(cats, str):
                cats = [c.strip() for c in cats.split(",") if c.strip()]
            if not isinstance(cats, list):
                cats = []
            cats = [str(c) for c in cats if c]
            if cats:
                invalid = [c for c in cats if c not in EXPORT_CATEGORIES]
                if invalid:
                    raise HTTPException(status_code=400, detail=f"Unknown export categories: {', '.join(invalid)}")
                include_map[reg] = sorted(set(cats))

            if isinstance(files, str):
                files = [f.strip() for f in files.split(",") if f.strip()]
            if isinstance(files, list) and files:
                sanitized = [Path(f).name for f in files if f]
                include_files_map[reg] = sorted(set(sanitized))

        if include_files_map:
            regions = sorted(include_files_map.keys())
            region = None
            include_map = None
        elif include_map:
            regions = sorted(include_map.keys())
            region = None

    if region:
        region = str(region).upper()
        if region not in list_regions():
            raise HTTPException(status_code=404, detail=f"Unknown region: {region}")
    if regions:
        regions = [str(r).upper() for r in regions if r]
        unknown = [r for r in regions if r not in list_regions()]
        if unknown:
            raise HTTPException(status_code=404, detail=f"Unknown regions: {', '.join(unknown)}")

    export_root = prepare_export_dir()
    stamp = dt.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    out_dir = export_root / f"export_{stamp}_{uuid.uuid4().hex[:8]}"
    out_dir.mkdir(parents=True, exist_ok=True)

    export_id = f"exp_{uuid.uuid4().hex}"
    requested_at = dt.datetime.utcnow().isoformat()
    conn = get_conn()
    try:
        ensure_export_table(conn)
        conn.execute(
            """
            INSERT INTO export_jobs (
              export_id, requested_at, region, status, requested_by
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (export_id, requested_at, region, "running", "app"),
        )
        conn.commit()
    finally:
        conn.close()

    try:
        run_excel_regeneration(
            out_dir,
            include=include or None,
            regions=regions,
            include_map=include_map,
            include_files_map=include_files_map,
        )
        zip_path = build_export_zip(out_dir, region=region, regions=regions)
    except HTTPException as exc:
        conn = get_conn()
        try:
            ensure_export_table(conn)
            conn.execute(
                """
                UPDATE export_jobs
                SET status = ?, completed_at = ?, error_message = ?
                WHERE export_id = ?
                """,
                ("failed", dt.datetime.utcnow().isoformat(), str(exc.detail), export_id),
            )
            conn.commit()
        finally:
            conn.close()
        raise
    except Exception as exc:
        conn = get_conn()
        try:
            ensure_export_table(conn)
            conn.execute(
                """
                UPDATE export_jobs
                SET status = ?, completed_at = ?, error_message = ?
                WHERE export_id = ?
                """,
                ("failed", dt.datetime.utcnow().isoformat(), str(exc), export_id),
            )
            conn.commit()
        finally:
            conn.close()
        raise

    conn = get_conn()
    try:
        ensure_export_table(conn)
        conn.execute(
            """
            UPDATE export_jobs
            SET status = ?, completed_at = ?, file_path = ?
            WHERE export_id = ?
            """,
            ("success", dt.datetime.utcnow().isoformat(), str(zip_path), export_id),
        )
        conn.commit()
    finally:
        conn.close()

    return FileResponse(
        zip_path,
        media_type="application/zip",
        filename=zip_path.name,
    )


@router.get("/reports/export_history")
def export_history(limit: int = 25):
    limit = max(1, min(int(limit), 200))
    conn = get_conn()
    try:
        ensure_export_table(conn)
        rows = conn.execute(
            """
            SELECT export_id, requested_at, completed_at, region, status, file_path, error_message, requested_by
            FROM export_jobs
            ORDER BY requested_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return JSONResponse({"rows": [row_to_dict(r) for r in rows]})
    finally:
        conn.close()


@router.get("/reports/export_download")
def export_download(export_id: str):
    conn = get_conn()
    try:
        ensure_export_table(conn)
        row = conn.execute(
            "SELECT file_path FROM export_jobs WHERE export_id = ?",
            (export_id,),
        ).fetchone()
        if not row or not row["file_path"]:
            raise HTTPException(status_code=404, detail="Export file not found")
        file_path = Path(row["file_path"]).resolve()
    finally:
        conn.close()

    export_root = prepare_export_dir().resolve()
    if export_root not in file_path.parents:
        raise HTTPException(status_code=400, detail="Invalid export file path")

    return FileResponse(file_path, media_type="application/zip", filename=file_path.name)


@router.get("/reports/output_library")
def output_library(run: str | None = None, limit: int = 200):
    limit = max(1, min(int(limit), 2000))
    runs = list_export_runs(limit=10)
    if not runs:
        return JSONResponse({"run": None, "runs": [], "rows": []})
    chosen = None
    if run:
        chosen = next((r for r in runs if r["id"] == run), None)
    if not chosen:
        chosen = runs[0]
    run_dir = Path(chosen["path"])
    manifest = load_regeneration_manifest(run_dir)
    rows: List[Dict[str, Any]] = []
    if manifest and manifest.get("regions"):
        for region in manifest["regions"]:
            region_id = region.get("region") or ""
            for item in region.get("files", []):
                rows.append({
                    "region": region_id,
                    "file": item.get("file"),
                    "category": item.get("category"),
                    "status": item.get("status"),
                    "generator": item.get("generator"),
                    "run": chosen["id"],
                    "run_at": chosen["updated_at"],
                })
    rows = rows[:limit]
    return JSONResponse({"run": chosen["id"], "runs": runs, "rows": rows})


@router.get("/reports/export_manifests")
def export_manifests(run: str | None = None):
    runs = list_export_runs(limit=10)
    if not runs:
        return JSONResponse({"run": None, "runs": [], "regions": []})
    chosen = None
    if run:
        chosen = next((r for r in runs if r["id"] == run), None)
    if not chosen:
        chosen = runs[0]
    run_dir = Path(chosen["path"])
    manifest = load_regeneration_manifest(run_dir)
    regions_out: List[Dict[str, Any]] = []
    if manifest and manifest.get("regions"):
        for region in manifest["regions"]:
            files = region.get("files", [])
            total = len(files)
            generated = sum(1 for f in files if f.get("status") == "generated")
            passthrough = sum(1 for f in files if f.get("status") == "passthrough")
            categories = sorted({f.get("category") for f in files if f.get("category")})
            regions_out.append({
                "region": region.get("region"),
                "total_files": total,
                "generated": generated,
                "passthrough": passthrough,
                "categories": categories,
                "run": chosen["id"],
                "run_at": chosen["updated_at"],
            })
    return JSONResponse({"run": chosen["id"], "runs": runs, "regions": regions_out})


@router.get("/reports/output_preview")
def output_preview(region: str, file: str, run: str | None = None):
    if not region or not file:
        raise HTTPException(status_code=400, detail="region and file are required")
    runs = list_export_runs(limit=10)
    if not runs:
        raise HTTPException(status_code=404, detail="No exports available")
    chosen = None
    if run:
        chosen = next((r for r in runs if r["id"] == run), None)
    if not chosen:
        chosen = runs[0]
    run_dir = Path(chosen["path"]).resolve()
    target = (run_dir / region.upper() / Path(file).name).resolve()
    if not target.exists():
        raise HTTPException(status_code=404, detail="Output file not found")
    if run_dir not in target.parents:
        raise HTTPException(status_code=400, detail="Invalid output path")
    preview = _preview_workbook(target)
    return JSONResponse({
        "run": chosen["id"],
        "region": region.upper(),
        "file": target.name,
        "files": [preview],
        "context": {"type": "output", "region": region.upper(), "file": target.name, "run": chosen["id"]},
    })


@router.get("/reports/output_sheet")
def output_sheet(region: str, file: str, sheet: str, run: str | None = None, offset: int = 0, limit: int = 50, q: str | None = None):
    if not region or not file or not sheet:
        raise HTTPException(status_code=400, detail="region, file, and sheet are required")
    runs = list_export_runs(limit=10)
    if not runs:
        raise HTTPException(status_code=404, detail="No exports available")
    chosen = None
    if run:
        chosen = next((r for r in runs if r["id"] == run), None)
    if not chosen:
        chosen = runs[0]
    run_dir = Path(chosen["path"]).resolve()
    target = (run_dir / region.upper() / Path(file).name).resolve()
    if not target.exists():
        raise HTTPException(status_code=404, detail="Output file not found")
    if run_dir not in target.parents:
        raise HTTPException(status_code=400, detail="Invalid output path")
    payload = _read_sheet_data(target, sheet, offset=offset, limit=limit, q=q)
    return JSONResponse({
        "run": chosen["id"],
        "region": region.upper(),
        "file": target.name,
        **payload,
    })
