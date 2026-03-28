from __future__ import annotations

import datetime as dt
import io
import json
import logging
import re
import uuid
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Dict, List

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import JSONResponse, Response

from app.core.config import ROOT, EXPORT_ROOT
from app.core.db import get_conn
from app.services.imports import list_input_files, new_job_dir, run_migration
from app.core.config import IMPORT_ROOT

router = APIRouter(prefix="/api")
logger = logging.getLogger("smm")

_executor = ThreadPoolExecutor(max_workers=1)

EXPORT_REGEN = {
    "individual_sales",
    "sku_summary",
    "township_summary",
    "van_wise_sku",
}

CORE_INPUTS = [
    {
        "kind": "table_daily",
        "label": "Table + DailySales",
        "required": True,
        "patterns": [
            "*Table and DailySales*.xlsx",
            "*Table*DailySales*.xlsx",
            "*DailySales*.xlsx",
        ],
    },
    {
        "kind": "pg_sales",
        "label": "PG Table + PGSales",
        "required": False,
        "patterns": [
            "*PG Sales*.xlsx",
            "*PG-Daily Sales*.xlsx",
        ],
    },
]


def normalize_filename(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", name.lower()).strip()


def classify_template_name(filename: str) -> str:
    key = normalize_filename(filename)
    if "individual" in key or "individual sale" in key:
        return "individual_sales"
    if "van wise" in key:
        return "van_wise_sku"
    if "sku summary" in key or "sku wise" in key or "sku analysis" in key:
        return "sku_summary"
    if "township" in key and "summary" in key:
        return "township_summary"
    if "sales compare" in key or key.startswith("compare"):
        return "sales_compare"
    if "follow up" in key or "followup" in key or "fus" in key:
        return "follow_up_sales"
    if "debtor" in key:
        return "debtors"
    return "other"


def core_input_kind(filename: str) -> tuple[str, str] | None:
    key = normalize_filename(filename)
    if "table" in key and ("dailysales" in key or "daily sales" in key):
        return "table_daily", "Table + DailySales"
    if "pg" in key and ("pg sales" in key or "pg daily sales" in key or "pg dailysales" in key):
        return "pg_sales", "PG Table + PGSales"
    return None


def core_input_reason(filename: str) -> str | None:
    kind = core_input_kind(filename)
    return kind[1] if kind else None


def build_source_catalog() -> Dict:
    source_dir = ROOT / "source"
    regions = []
    if not source_dir.exists():
        return {"regions": regions, "core_expected": CORE_INPUTS}
    for region_dir in sorted([p for p in source_dir.iterdir() if p.is_dir()]):
        core_files = []
        export_files = []
        files = sorted(
            [p for p in region_dir.iterdir() if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"}]
        )
        for f in files:
            core = core_input_kind(f.name)
            if core:
                kind, label = core
                core_files.append({"file": f.name, "kind": kind, "reason": label})
                continue
            category = classify_template_name(f.name)
            export_files.append({
                "file": f.name,
                "category": category,
                "export_type": "regen" if category in EXPORT_REGEN else "template",
            })
        expected_kinds = {c["kind"] for c in CORE_INPUTS}
        found_kinds = {c["kind"] for c in core_files if c.get("kind")}
        missing = sorted(expected_kinds - found_kinds)
        regions.append({
            "region": region_dir.name.upper(),
            "core_expected": CORE_INPUTS,
            "core": core_files,
            "missing_core": missing,
            "exports": export_files,
        })
    return {"regions": regions, "core_expected": CORE_INPUTS}


def _cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (dt.date, dt.datetime)):
        return val.isoformat()
    return str(val).strip()


def _preview_workbook_bytes(payload: bytes, name: str) -> Dict:
    wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        sheets = []
        for ws in wb.worksheets:
            header = []
            header_row_index = None
            for idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
                if not row:
                    continue
                if any(cell not in (None, "") for cell in row):
                    header = [_cell_to_str(cell) for cell in row]
                    header_row_index = idx
                    break
            while header and header[-1] == "":
                header.pop()
            sample_rows = []
            max_cols = min(max(len(header), 6), 12)
            start_row = header_row_index + 1 if header_row_index else 1
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + 4, values_only=True):
                if not row:
                    continue
                sample_rows.append([_cell_to_str(cell) for cell in row[:max_cols]])
            sheets.append({
                "name": ws.title,
                "rows": ws.max_row or 0,
                "cols": ws.max_column or 0,
                "header": header,
                "sample_rows": sample_rows,
            })
    finally:
        wb.close()
    return {"file": name, "sheets": sheets}


def _preview_workbook(path: Path) -> Dict:
    return _preview_workbook_bytes(path.read_bytes(), path.name)


def _read_sheet_data(path: Path, sheet: str, offset: int = 0, limit: int = 50, q: str | None = None) -> Dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet not found")
        ws = wb[sheet]
        header = []
        header_row_index = None
        for idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
            if not row:
                continue
            if any(cell not in (None, "") for cell in row):
                header = [_cell_to_str(cell) for cell in row]
                header_row_index = idx
                break
        while header and header[-1] == "":
            header.pop()
        start_row = header_row_index + 1 if header_row_index else 1
        rows = []
        total = 0
        query = (q or "").lower().strip()
        if not query:
            total = max((ws.max_row or 0) - start_row + 1, 0)
            for idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=0):
                if idx < offset:
                    continue
                if len(rows) >= limit:
                    break
                rows.append([_cell_to_str(cell) for cell in row])
        else:
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not row:
                    continue
                row_vals = [_cell_to_str(cell) for cell in row]
                hay = " ".join(row_vals).lower()
                if query not in hay:
                    continue
                if total >= offset and len(rows) < limit:
                    rows.append(row_vals)
                total += 1
        return {
            "sheet": sheet,
            "header": header,
            "rows": rows,
            "total": total,
            "offset": offset,
            "limit": limit,
        }
    finally:
        wb.close()


def ensure_import_table(conn) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS import_jobs (
          import_id TEXT PRIMARY KEY,
          requested_at TEXT NOT NULL,
          completed_at TEXT,
          region TEXT,
          status TEXT NOT NULL,
          input_dir TEXT,
          file_names TEXT,
          error_message TEXT,
          requested_by TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS ix_import_jobs_requested_at
          ON import_jobs (requested_at)
        """
    )


def list_export_runs(limit: int = 10) -> List[Dict]:
    runs: List[Dict] = []
    if not EXPORT_ROOT.exists():
        return runs
    candidates = [p for p in EXPORT_ROOT.iterdir() if p.is_dir() and p.name.startswith("export_")]
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    for p in candidates[:limit]:
        runs.append({
            "id": p.name,
            "path": str(p),
            "updated_at": dt.datetime.utcfromtimestamp(p.stat().st_mtime).isoformat(),
        })
    return runs


def load_regeneration_manifest(run_dir: Path) -> Dict | None:
    top_manifest = run_dir / "regeneration_manifest.json"
    if top_manifest.exists():
        try:
            return json.loads(top_manifest.read_text(encoding="utf-8"))
        except Exception:
            return None
    regions = []
    for region_dir in sorted([p for p in run_dir.iterdir() if p.is_dir() and p.name.isupper()]):
        manifest_path = region_dir / "_regeneration_manifest.json"
        if not manifest_path.exists():
            continue
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        regions.append(manifest)
    if regions:
        return {"regions": regions}
    return None


def deps_for_output(filename: str) -> List[str]:
    deps = ["table_daily"]
    key = normalize_filename(filename)
    if "pg" in key:
        deps.append("pg_sales")
    return deps


def build_dependency_manifest() -> Dict:
    catalog = build_source_catalog()
    conn = get_conn()
    import_jobs = []
    try:
        ensure_import_table(conn)
        import_jobs = conn.execute(
            """
            SELECT import_id, region, status, requested_at, completed_at
            FROM import_jobs
            ORDER BY requested_at DESC
            """,
        ).fetchall()
    finally:
        conn.close()

    latest_import: Dict[str, str] = {}
    for row in import_jobs:
        region = (row["region"] or "").upper()
        if not region:
            continue
        if row["status"] != "success":
            continue
        stamp = row["completed_at"] or row["requested_at"]
        if not stamp:
            continue
        if region not in latest_import or stamp > latest_import[region]:
            latest_import[region] = stamp

    runs = list_export_runs(limit=10)
    export_files_by_region: Dict[str, Dict[str, Dict]] = {}
    latest_export: Dict[str, str] = {}
    if runs:
        chosen = runs[0]
        run_dir = Path(chosen["path"])
        manifest = load_regeneration_manifest(run_dir)
        if manifest and manifest.get("regions"):
            for region in manifest["regions"]:
                region_id = (region.get("region") or "").upper()
                if not region_id:
                    continue
                latest_export[region_id] = chosen["updated_at"]
                export_files_by_region.setdefault(region_id, {})
                for item in region.get("files", []):
                    export_files_by_region[region_id][item.get("file")] = item

    regions_out = []
    for region in catalog.get("regions", []):
        region_id = region.get("region")
        core_expected = catalog.get("core_expected", [])
        core_present = region.get("core", [])
        present_kinds = {c.get("kind") for c in core_present if c.get("kind")}
        core_list = []
        for core in core_expected:
            core_list.append({
                "kind": core.get("kind"),
                "label": core.get("label"),
                "required": core.get("required", False),
                "patterns": core.get("patterns", []),
                "present": core.get("kind") in present_kinds,
                "files": [c.get("file") for c in core_present if c.get("kind") == core.get("kind")],
            })

        outputs = []
        outputs_ready = 0
        outputs_blocked = 0
        outputs_fresh = 0
        outputs_stale = 0
        outputs_not = 0
        for item in region.get("exports", []):
            deps = deps_for_output(item.get("file", ""))
            missing = [d for d in deps if d not in present_kinds]
            export_info = export_files_by_region.get(region_id, {}).get(item.get("file"))
            status = "not_exported"
            if missing:
                status = "blocked_missing_core"
                outputs_blocked += 1
            elif export_info:
                last_import = latest_import.get(region_id)
                last_export = latest_export.get(region_id)
                if last_import and last_export and last_export < last_import:
                    status = "exported_stale"
                    outputs_stale += 1
                else:
                    status = "exported_fresh"
                    outputs_fresh += 1
            else:
                outputs_not += 1

            if status in {"exported_fresh", "exported_stale", "not_exported"}:
                outputs_ready += 1

            outputs.append({
                "file": item.get("file"),
                "category": item.get("category"),
                "export_type": item.get("export_type"),
                "dependencies": deps,
                "missing_deps": missing,
                "status": status,
                "export_info": export_info,
            })

        regions_out.append({
            "region": region_id,
            "last_import_at": latest_import.get(region_id),
            "last_export_at": latest_export.get(region_id),
            "core": core_list,
            "outputs": outputs,
            "summary": {
                "core_missing": sum(1 for c in core_list if c.get("required") and not c.get("present")),
                "outputs_total": len(outputs),
                "outputs_blocked": outputs_blocked,
                "outputs_ready": outputs_ready,
                "outputs_exported_fresh": outputs_fresh,
                "outputs_exported_stale": outputs_stale,
                "outputs_not_exported": outputs_not,
            },
        })

    return {
        "generated_at": dt.datetime.utcnow().isoformat(),
        "regions": regions_out,
        "core_expected": catalog.get("core_expected", []),
    }


def update_job(import_id: str, **fields) -> None:
    conn = get_conn()
    try:
        ensure_import_table(conn)
        sets = []
        vals = []
        for k, v in fields.items():
            sets.append(f"{k} = ?")
            vals.append(v)
        vals.append(import_id)
        conn.execute(
            f"UPDATE import_jobs SET {', '.join(sets)} WHERE import_id = ?",
            vals,
        )
        conn.commit()
    finally:
        conn.close()


def run_import_job(import_id: str, in_dir: Path, region: str | None) -> None:
    try:
        run_migration(in_dir, region)
        update_job(
            import_id,
            status="success",
            completed_at=dt.datetime.utcnow().isoformat(),
        )
    except HTTPException as exc:
        update_job(
            import_id,
            status="failed",
            completed_at=dt.datetime.utcnow().isoformat(),
            error_message=str(exc.detail),
        )
    except Exception as exc:
        logger.exception("Import failed")
        update_job(
            import_id,
            status="failed",
            completed_at=dt.datetime.utcnow().isoformat(),
            error_message=str(exc),
        )


@router.post("/imports")
async def create_import(
    files: List[UploadFile] = File(...),
    region: str | None = Form(None),
):
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")
    if region:
        region = region.upper()

    job_dir = new_job_dir()
    saved = []
    for f in files:
        name = Path(f.filename).name
        dest = job_dir / name
        content = await f.read()
        dest.write_bytes(content)
        saved.append(name)

    import_id = f"imp_{uuid.uuid4().hex}"
    requested_at = dt.datetime.utcnow().isoformat()

    conn = get_conn()
    try:
        ensure_import_table(conn)
        conn.execute(
            """
            INSERT INTO import_jobs (
              import_id, requested_at, region, status, input_dir, file_names, requested_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                import_id,
                requested_at,
                region,
                "running",
                str(job_dir),
                json.dumps(saved),
                "app",
            ),
        )
        conn.commit()
    finally:
        conn.close()

    _executor.submit(run_import_job, import_id, job_dir, region)

    return JSONResponse({"import_id": import_id, "status": "running", "files": saved})


@router.get("/imports")
def list_imports(limit: int = 25):
    limit = max(1, min(int(limit), 200))
    conn = get_conn()
    try:
        ensure_import_table(conn)
        rows = conn.execute(
            """
            SELECT import_id, requested_at, completed_at, region, status, input_dir, file_names, error_message, requested_by
            FROM import_jobs
            ORDER BY requested_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        out = []
        for r in rows:
            d = dict(r)
            try:
                d["file_names"] = json.loads(d.get("file_names") or "[]")
            except Exception:
                d["file_names"] = []
            out.append(d)
        return JSONResponse({"rows": out})
    finally:
        conn.close()


@router.get("/imports/errors")
def import_errors(limit: int = 200, q: str | None = None):
    limit = max(1, min(int(limit), 1000))
    conn = get_conn()
    try:
        ensure_import_table(conn)
        rows = conn.execute(
            """
            SELECT import_id, requested_at, completed_at, region, status, file_names, error_message
            FROM import_jobs
            WHERE status = 'failed'
            ORDER BY requested_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        out = []
        query = (q or "").lower().strip()
        for r in rows:
            d = dict(r)
            try:
                d["file_names"] = json.loads(d.get("file_names") or "[]")
            except Exception:
                d["file_names"] = []
            if query:
                hay = f"{d.get('region','')} {d.get('error_message','')} {' '.join(d['file_names'])}".lower()
                if query not in hay:
                    continue
            out.append(d)
        return JSONResponse({"rows": out})
    finally:
        conn.close()


@router.get("/imports/catalog")
def imports_catalog():
    return JSONResponse(build_source_catalog())


@router.get("/imports/dependency_manifest")
def imports_dependency_manifest(download: int | None = None):
    manifest = build_dependency_manifest()
    if download:
        payload = json.dumps(manifest, ensure_ascii=False, indent=2)
        return Response(
            content=payload,
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=\"dependency_manifest.json\""},
        )
    return JSONResponse(manifest)


@router.get("/imports/file_preview")
def imports_file_preview(region: str, category: str):
    if not region or not category:
        raise HTTPException(status_code=400, detail="region and category are required")
    region = region.upper()
    source_dir = ROOT / "source" / region
    if not source_dir.exists():
        raise HTTPException(status_code=404, detail=f"Region {region} not found")
    files = [
        p for p in source_dir.iterdir()
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"}
    ]
    previews = []
    for f in sorted(files):
        if core_input_reason(f.name):
            continue
        if classify_template_name(f.name) != category:
            continue
        previews.append(_preview_workbook(f))
    return JSONResponse({"region": region, "category": category, "files": previews})


@router.get("/imports/source_preview")
def imports_source_preview(region: str, file: str):
    if not region or not file:
        raise HTTPException(status_code=400, detail="region and file are required")
    region = region.upper()
    source_dir = ROOT / "source" / region
    if not source_dir.exists():
        raise HTTPException(status_code=404, detail=f"Region {region} not found")
    filename = Path(file).name
    target = source_dir / filename
    if not target.exists():
        raise HTTPException(status_code=404, detail="File not found in source")
    preview = _preview_workbook(target)
    return JSONResponse({
        "region": region,
        "file": filename,
        "files": [preview],
        "context": {"type": "source", "region": region, "file": filename},
    })


@router.get("/imports/saved_preview")
def imports_saved_preview(import_id: str, file: str):
    if not import_id or not file:
        raise HTTPException(status_code=400, detail="import_id and file are required")
    conn = get_conn()
    try:
        ensure_import_table(conn)
        row = conn.execute(
            "SELECT input_dir FROM import_jobs WHERE import_id = ?",
            (import_id,),
        ).fetchone()
        if not row or not row["input_dir"]:
            raise HTTPException(status_code=404, detail="Import job not found")
        input_dir = Path(row["input_dir"]).resolve()
    finally:
        conn.close()

    import_root = IMPORT_ROOT.resolve()
    if import_root not in input_dir.parents and input_dir != import_root:
        raise HTTPException(status_code=400, detail="Invalid import directory")
    target = (input_dir / Path(file).name).resolve()
    if not target.exists():
        raise HTTPException(status_code=404, detail="Imported file not found")
    preview = _preview_workbook(target)
    return JSONResponse({
        "import_id": import_id,
        "file": target.name,
        "files": [preview],
        "context": {"type": "saved", "import_id": import_id, "file": target.name},
    })


@router.post("/imports/preview")
async def imports_preview(file: UploadFile = File(...)):
    if not file:
        raise HTTPException(status_code=400, detail="file is required")
    name = Path(file.filename or "upload.xlsx").name
    payload = await file.read()
    preview = _preview_workbook_bytes(payload, name)
    return JSONResponse({"files": [preview], "context": {"type": "upload"}})


@router.get("/imports/source_sheet")
def imports_source_sheet(region: str, file: str, sheet: str, offset: int = 0, limit: int = 50, q: str | None = None):
    if not region or not file or not sheet:
        raise HTTPException(status_code=400, detail="region, file, and sheet are required")
    region = region.upper()
    source_dir = ROOT / "source" / region
    if not source_dir.exists():
        raise HTTPException(status_code=404, detail=f"Region {region} not found")
    filename = Path(file).name
    target = source_dir / filename
    if not target.exists():
        raise HTTPException(status_code=404, detail="File not found in source")
    payload = _read_sheet_data(target, sheet, offset=offset, limit=limit, q=q)
    return JSONResponse({"region": region, "file": filename, **payload})


@router.get("/imports/saved_sheet")
def imports_saved_sheet(import_id: str, file: str, sheet: str, offset: int = 0, limit: int = 50, q: str | None = None):
    if not import_id or not file or not sheet:
        raise HTTPException(status_code=400, detail="import_id, file, and sheet are required")
    conn = get_conn()
    try:
        ensure_import_table(conn)
        row = conn.execute(
            "SELECT input_dir FROM import_jobs WHERE import_id = ?",
            (import_id,),
        ).fetchone()
        if not row or not row["input_dir"]:
            raise HTTPException(status_code=404, detail="Import job not found")
        input_dir = Path(row["input_dir"]).resolve()
    finally:
        conn.close()

    import_root = IMPORT_ROOT.resolve()
    if import_root not in input_dir.parents and input_dir != import_root:
        raise HTTPException(status_code=400, detail="Invalid import directory")
    target = (input_dir / Path(file).name).resolve()
    if not target.exists():
        raise HTTPException(status_code=404, detail="Imported file not found")
    payload = _read_sheet_data(target, sheet, offset=offset, limit=limit, q=q)
    return JSONResponse({"import_id": import_id, "file": target.name, **payload})


@router.get("/imports/{import_id}")
def import_detail(import_id: str):
    conn = get_conn()
    try:
        ensure_import_table(conn)
        row = conn.execute(
            "SELECT * FROM import_jobs WHERE import_id = ?",
            (import_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Import job not found")
        d = dict(row)
        try:
            d["file_names"] = json.loads(d.get("file_names") or "[]")
        except Exception:
            d["file_names"] = []
        if d.get("input_dir"):
            d["files_present"] = list_input_files(Path(d["input_dir"]))
        return JSONResponse(d)
    finally:
        conn.close()
