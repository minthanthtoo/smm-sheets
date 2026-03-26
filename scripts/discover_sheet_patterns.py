#!/usr/bin/env python3
"""Scan Excel workbooks to detect table patterns and headers.

Outputs a JSON report with detected sheet patterns and header matches.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List

import openpyxl

from etl_load_sources import (
    DAILY_HEADERS,
    OUTLET_HEADERS,
    PRODUCT_HEADERS,
    WAY_PLAN_HEADERS,
    norm_key,
)


PG_MARKERS = {"pg name", "pgname", "salesbot", "bonus", "mmk", "column2"}
MEASURE_TOKENS = {"pkt", "pk", "pack", "bot", "bottle", "lit", "liter", "litre"}


def safe_row(ws, r: int, max_col: int) -> List[str]:
    try:
        vals = next(ws.iter_rows(min_row=r, max_row=r, max_col=max_col, values_only=True))
    except StopIteration:
        return []
    return [norm_key(v) for v in vals]


def best_match_row(ws, keys: set[str], max_scan: int = 20) -> Dict:
    best = {"score": 0, "row": None, "matches": []}
    max_col = ws.max_column or 0
    for r in range(1, min(max_scan, ws.max_row or 0) + 1):
        row_vals = safe_row(ws, r, max_col)
        matches = sorted({v for v in row_vals if v in keys})
        if len(matches) > best["score"]:
            best = {"score": len(matches), "row": r, "matches": matches}
    return best


def detect_month_table(ws) -> Dict:
    max_scan = min(8, ws.max_row or 0)
    month_row = None
    measure_row = None
    max_col = ws.max_column or 0
    for r in range(1, max_scan + 1):
        row_vals = safe_row(ws, r, max_col)
        if any(v in MEASURE_TOKENS for v in row_vals):
            measure_row = r
        # month labels often appear as month names or numbers in the row
        if any(v in ("jan", "january", "feb", "february", "mar", "march", "apr", "april") for v in row_vals):
            month_row = r
    return {"month_row": month_row, "measure_row": measure_row, "score": 1 if month_row and measure_row else 0}


def classify_sheet(ws) -> Dict:
    max_col = ws.max_column or 0
    daily = best_match_row(ws, set(DAILY_HEADERS.keys()), max_scan=30)
    outlet = best_match_row(ws, set(OUTLET_HEADERS), max_scan=30)
    way = best_match_row(ws, set(WAY_PLAN_HEADERS), max_scan=30)
    row1 = safe_row(ws, 1, max_col)
    table_score = sum(1 for v in row1 if v in PRODUCT_HEADERS)
    pg_score = len([v for v in daily["matches"] if v in PG_MARKERS])
    month_table = detect_month_table(ws)

    patterns = []
    if daily["score"] >= 5:
        patterns.append("daily")
    if pg_score >= 2:
        patterns.append("daily_pg")
    if outlet["score"] >= 4:
        patterns.append("outlet_list")
    if way["score"] >= 5:
        patterns.append("way_plan")
    if table_score >= 3:
        patterns.append("table")
    if month_table["score"] >= 1:
        patterns.append("summary_monthly")
    if not patterns:
        patterns.append("unknown")

    return {
        "patterns": patterns,
        "daily": daily,
        "outlet": outlet,
        "way_plan": way,
        "table_score": table_score,
        "pg_score": pg_score,
        "month_table": month_table,
    }


def scan_paths(paths: List[Path]) -> List[Dict]:
    report = []
    for path in paths:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True, keep_links=False)
        for ws in wb.worksheets:
            info = classify_sheet(ws)
            report.append({
                "file": path.name,
                "sheet": ws.title,
                **info,
            })
        wb.close()
    return report


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=".")
    parser.add_argument("--out", default="")
    args = parser.parse_args()

    root = Path(args.root).resolve()
    candidates = []
    for sub in ("in", "source", "files"):
        base = root / sub
        if base.exists():
            candidates.extend(base.rglob("*.xlsx"))
            candidates.extend(base.rglob("*.xlsm"))
    candidates = sorted(set(candidates))

    report = scan_paths(candidates)
    payload = {
        "files_scanned": len(candidates),
        "sheets_scanned": len(report),
        "records": report,
    }

    if args.out:
        out_path = Path(args.out).resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    else:
        print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
